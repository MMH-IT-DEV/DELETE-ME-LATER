"""
Generates the proposed GMP-compliant Systems Health Tracker XLSX.
Run: python3 build-health-tracker.py
Output: health-tracker-proposal.xlsx
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

wb = openpyxl.Workbook()

# ─────────────────────────────────────────────
# COLOURS
# ─────────────────────────────────────────────
C = {
    "header_bg":   "263238", "header_fg":   "FFFFFF",
    "sep_flow":    "01579B", "sep_script":  "004D40",
    "sep_bot":     "4A148C", "sep_sheet":   "1B5E20",
    "row_even":    "F5F7F9", "row_odd":     "FFFFFF",
    "plat_shopify":"00897B", "plat_google": "1565C0",
    "plat_fedex":  "6A1B9A", "plat_katana": "E65100",
    "plat_wasp":   "2E7D32",
    "ok_bg":       "0D5016", "ok_fg":       "4ADE80",
    "warn_bg":     "553600", "warn_fg":     "FBBF24",
    "down_bg":     "5C0011", "down_fg":     "F87171",
    "unk_bg":      "3B3B3B", "unk_fg":      "D1D5DB",
    "pend_bg":     "1A237E", "pend_fg":     "90CAF9",
    "gmp_yes_bg":  "1B5E20", "gmp_yes_fg":  "A5D6A7",
    "gmp_no_bg":   "37474F", "gmp_no_fg":   "ECEFF1",
    "expiry_warn": "553600", "expiry_ok":   "1B5E20",
    "hb_yes_bg":   "004D40", "hb_yes_fg":   "80CBC4",
    "days_warn":   "553600", "days_ok":     "1B5E20",
}

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(hex_color="000000", bold=False, size=10):
    return Font(color=hex_color, bold=bold, size=size, name="Segoe UI")

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=False)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=False)

def thin_border():
    s = Side(style="thin", color="CFD8DC")
    return Border(bottom=s)

# ─────────────────────────────────────────────
# TAB 1 — SYSTEM REGISTRY
# ─────────────────────────────────────────────
ws1 = wb.active
ws1.title = "System Registry"
ws1.freeze_panes = "A2"
ws1.row_dimensions[1].height = 22

HEADERS = [
    ("System Name",       32),
    ("Description",       40),
    ("Type",              13),
    ("Platform",          13),
    ("GMP Critical",      13),
    ("Run Frequency",     16),
    ("Maintenance Guide", 26),
    ("Auth Type",         15),
    ("Auth Location",     28),
    ("Expiry Date",       13),
    ("Days Left",         10),
    ("Owner",             12),
    ("Validated",         12),
    ("Last Validated",    15),
    ("Heartbeat Method",  20),
    ("Status",            13),
    ("Last Heartbeat",    20),
    ("Last Run OK",       12),
    ("Notes",             28),
]

# Write headers
for col_idx, (label, width) in enumerate(HEADERS, start=1):
    cell = ws1.cell(row=1, column=col_idx, value=label)
    cell.fill      = fill(C["header_bg"])
    cell.font      = font(C["header_fg"], bold=True, size=10)
    cell.alignment = center()
    ws1.column_dimensions[get_column_letter(col_idx)].width = width

# ── System data ──────────────────────────────
SYSTEMS = {
    "FLOW": [
        ["Shopify Hold → ShipStation Hold Sync",
         "Puts ShipStation order on hold when Shopify order is held",
         "FLOW","Shopify","Yes","Per order",
         "SOP-008-ShipStation-Shopify-Flow",
         "Bearer Token","Shopify Secrets → Secrets Manager",
         "2027-01-01", 304,"Felippe",
         "Pending","","Flow webhook step","Unknown","","",""],

        ["Shopify Hold Release → ShipStation Release",
         "Releases ShipStation hold when Shopify hold is removed",
         "FLOW","Shopify","Yes","Per order",
         "SOP-008-ShipStation-Shopify-Flow",
         "Bearer Token","Shopify Secrets → Secrets Manager",
         "2027-01-01", 305,"Felippe",
         "Pending","","Flow webhook step","Unknown","","",""],

        ["Shopify Cancel → ShipStation Cancel",
         "Cancels ShipStation order when Shopify order is cancelled",
         "FLOW","Shopify","Yes","Per order",
         "SOP-008-ShipStation-Shopify-Flow",
         "Bearer Token","Shopify Secrets → Secrets Manager",
         "2027-01-01", 305,"Felippe",
         "Pending","","Flow webhook step","Unknown","","",""],
    ],
    "SCRIPT": [
        ["Katana → Google Calendar PO Sync",
         "Syncs Katana purchase orders to Google Calendar as dated events",
         "SCRIPT","Katana","Yes","Per PO event",
         "SOP-005-Katana-Calendar-PO-Sync",
         "API Key","Apps Script → Script Properties",
         "NO EXPIRY","","Erik",
         "Yes","2026-01-26","Script sendHeartbeat()","Healthy","2026-01-28 00:23","Yes",""],

        ["QA Escalation Automation",
         "Escalates QA issues via Google Sheet triggers and Slack alerts",
         "SCRIPT","Google","Yes","Daily + on edit",
         "SOP-007-QA-Escalation-Bot",
         "None","N/A",
         "","","Erik",
         "Pending","","Script sendHeartbeat()","Healthy","2026-01-27","Yes",""],
    ],
    "BOT": [
        ["FedEx Dispute Bot",
         "Automatically files FedEx shipping disputes from GitHub-hosted bot",
         "BOT","FedEx","No","Weekly",
         "SOP-006-FedEx-Dispute-Bot",
         "None","N/A",
         "","","Erik",
         "Pending","","Bot HTTP POST","Healthy","2026-03-02","Yes",""],
    ],
    "GOOGLE SHEET": [
        ["2026_Security & Connection Tracker",
         "GMP access log, incident register, and periodic review tracker",
         "GOOGLE SHEET","Google","Yes","On demand",
         "SOP-007-Security-Tracker",
         "None","N/A",
         "","","Erik",
         "Yes","2026-01-01","N/A","Healthy","N/A","N/A",""],

        ["WASP-KATANA-SYNC",
         "Inventory sync dashboard — Katana MRP ↔ WASP InventoryCloud",
         "GOOGLE SHEET","Wasp/Katana","Yes","Hourly",
         "SOP-WASP-Katana-Sync",
         "API Key","Katana Settings → Script Properties",
         "","","Erik",
         "Yes","2026-01-01","N/A","Healthy","N/A","N/A",""],
    ],
}

SEP_COLOR_MAP = {
    "FLOW":         C["sep_flow"],
    "SCRIPT":       C["sep_script"],
    "BOT":          C["sep_bot"],
    "GOOGLE SHEET": C["sep_sheet"],
}

PLATFORM_COLOR_MAP = {
    "SHOPIFY":     C["plat_shopify"],
    "GOOGLE":      C["plat_google"],
    "FEDEX":       C["plat_fedex"],
    "KATANA":      C["plat_katana"],
    "WASP/KATANA": C["plat_wasp"],
}

current_row = 2
TYPE_ORDER = ["FLOW", "SCRIPT", "BOT", "GOOGLE SHEET"]

for type_name in TYPE_ORDER:
    rows = SYSTEMS.get(type_name, [])
    if not rows:
        continue

    sep_color = SEP_COLOR_MAP[type_name]
    num_cols  = len(HEADERS)

    # Separator row
    ws1.row_dimensions[current_row].height = 16
    for c in range(1, num_cols + 1):
        cell = ws1.cell(row=current_row, column=c, value=(type_name if c == 2 else ""))
        cell.fill      = fill(sep_color)
        cell.font      = font("FFFFFF", bold=True, size=9)
        cell.alignment = left()
    current_row += 1

    for dr, row_data in enumerate(rows):
        ws1.row_dimensions[current_row].height = 18
        row_bg = C["row_even"] if dr % 2 == 0 else C["row_odd"]

        for c_idx, value in enumerate(row_data, start=1):
            cell = ws1.cell(row=current_row, column=c_idx, value=value)
            cell.fill      = fill(row_bg)
            cell.font      = font("212121", size=9)
            cell.alignment = left()
            cell.border    = thin_border()

        # Platform chip — col 4
        plat_key = str(row_data[3]).strip().upper()
        plat_col = PLATFORM_COLOR_MAP.get(plat_key)
        if plat_col:
            c = ws1.cell(row=current_row, column=4)
            c.fill = fill(plat_col)
            c.font = font("FFFFFF", bold=True, size=9)
            c.alignment = center()

        # GMP Critical — col 5
        gmp_val = str(row_data[4]).strip()
        gmp_cell = ws1.cell(row=current_row, column=5)
        if gmp_val == "Yes":
            gmp_cell.fill = fill(C["gmp_yes_bg"])
            gmp_cell.font = font(C["gmp_yes_fg"], bold=True, size=9)
        else:
            gmp_cell.fill = fill(C["gmp_no_bg"])
            gmp_cell.font = font(C["gmp_no_fg"], size=9)
        gmp_cell.alignment = center()

        # Validated — col 13
        val_val  = str(row_data[12]).strip()
        val_cell = ws1.cell(row=current_row, column=13)
        if val_val == "Yes":
            val_cell.fill = fill(C["ok_bg"])
            val_cell.font = font(C["ok_fg"], bold=True, size=9)
        elif val_val == "Pending":
            val_cell.fill = fill(C["pend_bg"])
            val_cell.font = font(C["pend_fg"], bold=True, size=9)
        val_cell.alignment = center()

        # Heartbeat Method — col 15
        hb_val  = str(row_data[14]).strip()
        hb_cell = ws1.cell(row=current_row, column=15)
        if hb_val != "N/A" and hb_val != "":
            hb_cell.fill = fill(C["hb_yes_bg"])
            hb_cell.font = font(C["hb_yes_fg"], size=9)
        hb_cell.alignment = center()

        # Status — col 16
        status   = str(row_data[15]).strip()
        s_cell   = ws1.cell(row=current_row, column=16)
        s_cell.alignment = center()
        if status == "Healthy":
            s_cell.fill = fill(C["ok_bg"])
            s_cell.font = font(C["ok_fg"], bold=True, size=9)
        elif status == "Degraded":
            s_cell.fill = fill(C["warn_bg"])
            s_cell.font = font(C["warn_fg"], bold=True, size=9)
        elif status == "Down":
            s_cell.fill = fill(C["down_bg"])
            s_cell.font = font(C["down_fg"], bold=True, size=9)
        else:
            s_cell.fill = fill(C["unk_bg"])
            s_cell.font = font(C["unk_fg"], bold=True, size=9)

        # Days Left — col 11
        days_val  = row_data[10]
        days_cell = ws1.cell(row=current_row, column=11)
        days_cell.alignment = center()
        if isinstance(days_val, int):
            if days_val <= 30:
                days_cell.fill = fill(C["warn_bg"])
                days_cell.font = font(C["warn_fg"], bold=True, size=9)
            else:
                days_cell.fill = fill(C["days_ok"])
                days_cell.font = font("A5D6A7", size=9)

        current_row += 1

# ─────────────────────────────────────────────
# TAB 2 — HEARTBEAT SETUP GUIDE
# ─────────────────────────────────────────────
ws2 = wb.create_sheet("Heartbeat Setup Guide")
ws2.freeze_panes = "A2"
ws2.row_dimensions[1].height = 22

HB_HEADERS = [
    ("System Name",        30),
    ("Type",               13),
    ("Heartbeat Method",   22),
    ("Integration Step",   50),
    ("Payload Example",    55),
    ("Threshold (hours)",  18),
    ("Setup Status",       15),
    ("Who Sets It Up",     18),
]

for col_idx, (label, width) in enumerate(HB_HEADERS, start=1):
    cell = ws2.cell(row=1, column=col_idx, value=label)
    cell.fill      = fill(C["header_bg"])
    cell.font      = font(C["header_fg"], bold=True, size=10)
    cell.alignment = center()
    ws2.column_dimensions[get_column_letter(col_idx)].width = width

HB_DATA = [
    # FLOW separator
    ("_SEP_", "FLOW"),
    ["Shopify Hold → ShipStation Hold Sync",
     "FLOW",
     "Flow webhook step",
     "Add 'Send HTTP Request' action at END of flow.\nURL = Health Monitor web app URL\nMethod = POST",
     '{"system":"Shopify Hold Sync","status":"success","trigger":"Order #{{order_id}}"}',
     48, "Pending", "Felippe"],

    ["Shopify Hold Release → ShipStation Release",
     "FLOW",
     "Flow webhook step",
     "Add 'Send HTTP Request' action at END of flow.\nURL = Health Monitor web app URL",
     '{"system":"Shopify Hold Release","status":"success","trigger":"Order #{{order_id}}"}',
     48, "Pending", "Felippe"],

    ["Shopify Cancel → ShipStation Cancel",
     "FLOW",
     "Flow webhook step",
     "Add 'Send HTTP Request' action at END of flow.\nURL = Health Monitor web app URL",
     '{"system":"Shopify Cancel Sync","status":"success","trigger":"Order #{{order_id}}"}',
     48, "Pending", "Felippe"],

    # SCRIPT separator
    ("_SEP_", "SCRIPT"),
    ["Katana → Google Calendar PO Sync",
     "SCRIPT",
     "Script sendHeartbeat()",
     "Add sendHeartbeatToMonitor_() call at end of main function in GAS script.",
     '{"system":"Katana Calendar PO Sync","status":"success","details":"X events created"}',
     48, "Ready to add", "Erik"],

    ["QA Escalation Automation",
     "SCRIPT",
     "Script sendHeartbeat()",
     "Add sendHeartbeatToMonitor_() call at end of dailyChecks() in GAS script.",
     '{"system":"QA Escalation Automation","status":"success","details":"X checks run"}',
     48, "Ready to add", "Erik"],

    # BOT separator
    ("_SEP_", "BOT"),
    ["FedEx Dispute Bot",
     "BOT",
     "Bot HTTP POST",
     "Add requests.post() call at end of main bot execution (GitHub repo).",
     '{"system":"FedEx Dispute Bot","status":"success","details":"X disputes filed"}',
     192, "Needs dev", "Erik / Felippe"],

    # GOOGLE SHEET separator
    ("_SEP_", "GOOGLE SHEET"),
    ["2026_Security & Connection Tracker",
     "GOOGLE SHEET",
     "N/A",
     "Sheet-only system. No active script to heartbeat.\nStatus updated manually or via daily health check.",
     "N/A",
     0, "N/A", "N/A"],

    ["WASP-KATANA-SYNC",
     "GOOGLE SHEET",
     "N/A",
     "Sheet-only system. No active script to heartbeat.\nStatus updated manually or via daily health check.",
     "N/A",
     0, "N/A", "N/A"],
]

hb_row = 2
for entry in HB_DATA:
    ws2.row_dimensions[hb_row].height = 42

    if isinstance(entry, tuple) and entry[0] == "_SEP_":
        type_name  = entry[1]
        sep_color  = SEP_COLOR_MAP.get(type_name, "424242")
        for c in range(1, len(HB_HEADERS) + 1):
            cell = ws2.cell(row=hb_row, column=c, value=(type_name if c == 2 else ""))
            cell.fill      = fill(sep_color)
            cell.font      = font("FFFFFF", bold=True, size=9)
            cell.alignment = left()
        hb_row += 1
        continue

    for c_idx, value in enumerate(entry, start=1):
        cell = ws2.cell(row=hb_row, column=c_idx, value=value)
        cell.fill      = fill(C["row_even"])
        cell.font      = font("212121", size=9)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.border    = thin_border()

    # Type chip — col 2
    type_chip = ws2.cell(row=hb_row, column=2)
    tcolor    = SEP_COLOR_MAP.get(str(entry[1]).strip(), "424242")
    type_chip.fill = fill(tcolor)
    type_chip.font = font("FFFFFF", bold=True, size=9)
    type_chip.alignment = center()

    # Setup Status — col 7
    status_val  = str(entry[6]).strip()
    status_cell = ws2.cell(row=hb_row, column=7)
    status_cell.alignment = center()
    if status_val == "N/A":
        status_cell.fill = fill(C["unk_bg"])
        status_cell.font = font(C["unk_fg"], size=9)
    elif status_val == "Ready to add":
        status_cell.fill = fill(C["ok_bg"])
        status_cell.font = font(C["ok_fg"], bold=True, size=9)
    elif status_val == "Pending":
        status_cell.fill = fill(C["pend_bg"])
        status_cell.font = font(C["pend_fg"], bold=True, size=9)
    elif status_val == "Needs dev":
        status_cell.fill = fill(C["warn_bg"])
        status_cell.font = font(C["warn_fg"], bold=True, size=9)

    hb_row += 1

# ─────────────────────────────────────────────
# SAVE
# ─────────────────────────────────────────────
output_path = Path(__file__).resolve().with_name("health-tracker-proposal.xlsx")
wb.save(output_path)
print("Saved: " + str(output_path))
