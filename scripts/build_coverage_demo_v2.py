"""
Flow Coverage Demo — matches real 2026-Katana-WASP_DebugLog styling exactly.
Adds the coverage panel to the Activity tab header without overlap.
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import copy, datetime

SRC  = r"C:\Users\Admin\Downloads\2026-Katana-WASP_DebugLog.xlsx"
OUT  = r"C:\Users\Admin\Documents\claude-projects\wasp-katana\docs\flow-coverage-demo.xlsx"

# ── Exact colours from the real sheet ───────────────────────────────────────
HDR_BG   = "1C2333"   # all 3 header rows
HDR_TXT  = "FFFFFF"
DIM_TXT  = "8899AA"   # secondary labels

# Flow row colours (match GAS 08_Logging.js)
FLOW_BG  = {
    "F1": "CCE5FF", "F2": "B2DFDB", "F3": "FCE4D6",
    "F4": "F0E6F6", "F5": "D1ECF1", "F6": "FFE0B2",
}
SUB_OK   = "E8F5E9"
SUB_WARN = "FFF8E1"
SUB_ERR  = "FFEBEE"
STATUS_OK  = "D4EDDA"
STATUS_ERR = "F8D7DA"
STATUS_SKP = "FFF8E1"

# Coverage panel colours
PANEL_HDR   = "111E2D"   # slightly darker than HDR_BG for contrast
PANEL_MATCH = "1B4332"   # green bg
PANEL_MATCH_TXT = "95D5B2"
PANEL_WARN  = "3D2000"   # amber bg
PANEL_WARN_TXT  = "FFB347"
PANEL_MISS  = "3D0000"   # red bg
PANEL_MISS_TXT  = "FF6B6B"
PANEL_NA    = "1C2333"   # same as header (grey / no-op)
PANEL_NA_TXT    = "556677"

def F(hex_): return PatternFill("solid", fgColor=hex_)
def fnt(bold=False, color=HDR_TXT, size=9, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")
def aln(h="center", v="center", wrap=False, indent=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)


def build_activity_sheet(wb, src_ws):
    """Clone Activity sheet then add the coverage panel."""
    ws = wb.create_sheet("Activity")

    # ── Copy column widths ────────────────────────────────────────────────
    for col, dim in src_ws.column_dimensions.items():
        ws.column_dimensions[col].width = dim.width
    for row, dim in src_ws.row_dimensions.items():
        ws.row_dimensions[row].height = dim.height

    # Add panel columns (G = thin spacer, H-M = 6 flow boxes)
    ws.column_dimensions["G"].width  = 2.5
    ws.column_dimensions["H"].width  = 10.5
    ws.column_dimensions["I"].width  = 10.5
    ws.column_dimensions["J"].width  = 10.5
    ws.column_dimensions["K"].width  = 10.5
    ws.column_dimensions["L"].width  = 10.5
    ws.column_dimensions["M"].width  = 10.5
    ws.column_dimensions["N"].width  = 12    # "Updated" cell
    ws.row_dimensions[1].height      = 22
    ws.row_dimensions[2].height      = 34
    ws.row_dimensions[3].height      = 15

    # ── Copy rows 1-3 (headers) from source ──────────────────────────────
    for r in range(1, 4):
        for c in range(1, 7):
            src = src_ws.cell(r, c)
            dst = ws.cell(r, c)
            dst.value = src.value
            dst.font  = copy.copy(src.font)
            dst.fill  = copy.copy(src.fill)
            dst.alignment = copy.copy(src.alignment)

    # Keep G1-N3 in same dark header colour
    for r in range(1, 4):
        for c in range(7, 15):
            ws.cell(r, c).fill = F(HDR_BG)

    ws.freeze_panes = "A4"

    # ── Coverage panel — Row 1: merged title bar ──────────────────────────
    ws.merge_cells("H1:M1")
    h1 = ws["H1"]
    h1.value     = "DAILY FLOW COVERAGE"
    h1.fill      = F(PANEL_HDR)
    h1.font      = fnt(bold=True, color=HDR_TXT, size=9)
    h1.alignment = aln()

    # ── Coverage panel — Row 2: one box per flow ─────────────────────────
    # Today's real numbers (from actual log + demo discrepancy)
    today_str = "2026-03-05"
    flows = [
        # col, label, name,           logged, expected
        ("H", "F1", "Receiving",      0,       0 ),   # no POs today → N/A
        ("I", "F2", "Adjustments",    6,       7 ),   # ← 1 missed (demo alert)
        ("J", "F3", "Transfers",      0,       0 ),   # no STs today → N/A
        ("K", "F4", "Manufacturing",  9,       9 ),
        ("L", "F5", "Shipping",      64,      64 ),
        ("M", "F6", "Amazon FBA",     0,       0 ),
    ]

    for col, lbl, name, logged, expected in flows:
        cell = ws[col + "2"]
        na = (expected == 0)

        if na:
            bg  = PANEL_NA;    txt = PANEL_NA_TXT;    mark = "–"
            ratio = "–"
        elif logged == expected:
            bg  = PANEL_MATCH; txt = PANEL_MATCH_TXT; mark = "✓"
            ratio = "{}/{}".format(logged, expected)
        elif logged > 0:
            bg  = PANEL_WARN;  txt = PANEL_WARN_TXT;  mark = "!"
            ratio = "{}/{}".format(logged, expected)
        else:
            bg  = PANEL_MISS;  txt = PANEL_MISS_TXT;  mark = "✗"
            ratio = "{}/{}".format(logged, expected)

        cell.value     = "{} {}\n{} {}".format(lbl, name, ratio, mark)
        cell.fill      = F(bg)
        cell.font      = Font(bold=True, color=txt, size=8, name="Calibri")
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)

        # Accent border only on active (non-NA) cells
        if not na:
            accent = "1DB954" if logged == expected else ("FFA726" if logged > 0 else "EF5350")
            s = Side(style="medium", color=accent)
            cell.border = Border(left=s, right=s, top=s, bottom=s)

    # "Last updated" cell — set value on N1 before merging N1:N2
    ws.cell(1, 14).value     = "Synced\n08:00\n{}".format(today_str)
    ws.cell(1, 14).fill      = F(PANEL_HDR)
    ws.cell(1, 14).font      = fnt(color=DIM_TXT, size=8)
    ws.cell(1, 14).alignment = aln(wrap=True)
    ws.merge_cells(start_row=1, start_column=14, end_row=2, end_column=14)

    # ── Copy sample data rows 4-60 ────────────────────────────────────────
    for r in range(4, min(61, src_ws.max_row + 1)):
        for c in range(1, 7):
            src = src_ws.cell(r, c)
            dst = ws.cell(r, c)
            dst.value     = src.value
            dst.font      = copy.copy(src.font)
            dst.fill      = copy.copy(src.fill)
            dst.alignment = copy.copy(src.alignment)
        ws.row_dimensions[r].height = 13

    # ── Add a fake "missed" F2 entry to illustrate the gap ───────────────
    # Insert a greyed-out placeholder row showing what was expected but not logged
    insert_row = 62
    ws.cell(insert_row, 1).value = "⚠ MISSED"
    ws.cell(insert_row, 1).font  = Font(bold=True, color="BB3300", size=9, name="Calibri")
    ws.cell(insert_row, 2).value = "2026-03-05 11:44"
    ws.cell(insert_row, 3).value = "F2 Adjustments"
    ws.cell(insert_row, 4).value = "Expected: WASP Adjustment  LCP-2 x1 — not found in log"
    ws.cell(insert_row, 5).value = "MISSING"
    for c in range(1, 7):
        ws.cell(insert_row, c).fill = F("FFF0F0")
        ws.cell(insert_row, c).font = Font(color="CC2200", size=9,
                                           italic=True, name="Calibri")
    ws.cell(insert_row, 5).font = Font(bold=True, color="CC2200", size=9, name="Calibri")

    return ws


def build_flow_tab(wb, src_ws, title):
    """Clone a flow tab (F1/F3/F4/F5) preserving headers + first 30 data rows."""
    ws = wb.create_sheet(title)
    for col, dim in src_ws.column_dimensions.items():
        ws.column_dimensions[col].width = dim.width
    for row, dim in src_ws.row_dimensions.items():
        ws.row_dimensions[row].height = dim.height

    for r in range(1, min(34, src_ws.max_row + 1)):
        for c in range(1, src_ws.max_column + 1):
            src = src_ws.cell(r, c)
            dst = ws.cell(r, c)
            dst.value     = src.value
            dst.font      = copy.copy(src.font)
            dst.fill      = copy.copy(src.fill)
            dst.alignment = copy.copy(src.alignment)
    ws.freeze_panes = "A4"
    return ws


def build():
    src_wb = openpyxl.load_workbook(SRC)
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # 1. Activity (with coverage panel)
    build_activity_sheet(wb, src_wb["Activity"])

    # 2. Flow tabs — clone first 30 data rows each
    for tab in ["F2 Adjustments", "F1 Receiving", "F3 Transfers",
                "F4 Manufacturing", "F5 Shipping"]:
        if tab in src_wb.sheetnames:
            build_flow_tab(wb, src_wb[tab], tab)

    wb.save(OUT)
    print("Saved:", OUT)


if __name__ == "__main__":
    build()
