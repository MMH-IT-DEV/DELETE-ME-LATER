"""
Build demo XLSX: Flow Coverage Counter in Activity Log
Shows how the daily F1/F2/F3/F5 counters sit in the header
without overlapping existing content.
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import os

# ── Palette ────────────────────────────────────────────────────────────────
DARK_NAVY   = "0D1B2A"
MID_NAVY    = "1A2E44"
LIGHT_NAVY  = "1F3A5C"
WHITE       = "FFFFFF"
GREY_TEXT   = "AAAAAA"

# Flow colours (match existing GAS activity log)
F1_BG  = "CCE5FF"   # blue-ish
F2_BG  = "B2DFDB"   # teal
F3_BG  = "FCE4D6"   # orange
F4_BG  = "F0E6F6"   # purple
F5_BG  = "D1ECF1"   # cyan

GREEN_OK    = "1B5E20"   # dark green text
GREEN_BG    = "E8F5E9"   # light green cell
AMBER_BG    = "FFF8E1"   # light amber cell
AMBER_TEXT  = "E65100"
RED_BG      = "FFEBEE"
RED_TEXT     = "B71C1C"
PANEL_BG    = "0F2640"   # counter panel background
PANEL_HDR   = "143454"   # counter panel header
PANEL_MATCH = "1B4332"   # green bg for matched count
PANEL_WARN  = "7B4F00"   # amber bg for partial
PANEL_MISS  = "7B1515"   # red bg for miss

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=WHITE, size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Calibri")

def center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=False)

def thin_border(sides="all"):
    s = Side(style="thin", color="2C2C2C")
    n = Side(style=None)
    if sides == "all":
        return Border(left=s, right=s, top=s, bottom=s)
    b = Border()
    if "l" in sides: b.left = s
    if "r" in sides: b.right = s
    if "t" in sides: b.top = s
    if "b" in sides: b.bottom = s
    return b


def build(path):
    wb = openpyxl.Workbook()

    # ── Sheet 1: Activity Log ─────────────────────────────────────────────
    ws = wb.active
    ws.title = "Activity"

    # Column widths (matching GAS sheet)
    col_widths = {
        "A": 10,   # ID
        "B": 14,   # Time
        "C": 16,   # Flow
        "D": 62,   # Details
        "E": 10,   # Status
        "F": 22,   # Error
        "G": 5,    # Retry (narrow spacer)
        "H": 9,    # ── PANEL ──
        "I": 9,
        "J": 9,
        "K": 9,
        "L": 15,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 15

    # ── Row 1: Title bar ─────────────────────────────────────────────────
    for c in range(1, 13):
        ws.cell(1, c).fill = fill(DARK_NAVY)

    ws["A1"] = "ACTIVITY LOG"
    ws["A1"].font = Font(bold=True, color=WHITE, size=13, name="Calibri")
    ws["A1"].alignment = left()

    ws["D1"] = "Katana-WASP Inventory Sync"
    ws["D1"].font = Font(bold=False, color=GREY_TEXT, size=10, name="Calibri", italic=True)
    ws["D1"].alignment = left()

    # ── Row 2: Sub-title bar ──────────────────────────────────────────────
    for c in range(1, 13):
        ws.cell(2, c).fill = fill(MID_NAVY)

    ws["A2"] = "All Flows — Chronological"
    ws["A2"].font = Font(bold=False, color=GREY_TEXT, size=9, name="Calibri", italic=True)
    ws["A2"].alignment = left()

    # ── FLOW COVERAGE PANEL  (H1:L2) ─────────────────────────────────────
    #
    #   H1:L1  merged → "DAILY FLOW COVERAGE  |  2026-03-05"
    #
    #   H2     I2      J2      K2      L2
    #   F1     F2      F3      F5      Updated
    #   3/3✓   12/12✓  2/2✓   47/47✓  08:00 ✓
    #
    # Row 1 panel header (merged H1:L1)
    ws.merge_cells("H1:L1")
    ws["H1"] = "DAILY FLOW COVERAGE  ·  2026-03-05"
    ws["H1"].fill = fill(PANEL_HDR)
    ws["H1"].font = Font(bold=True, color=WHITE, size=9, name="Calibri")
    ws["H1"].alignment = center()

    # Row 2: flow labels
    today_str = datetime.today().strftime("%Y-%m-%d")
    flows = [
        {"col": "H", "label": "F1",    "count": 3,  "expected": 3,  "name": "Receiving"},
        {"col": "I", "label": "F2",    "count": 12, "expected": 12, "name": "Adjustments"},
        {"col": "J", "label": "F3",    "count": 2,  "expected": 2,  "name": "Transfers"},
        {"col": "K", "label": "F5",    "count": 45, "expected": 47, "name": "Fulfillment"},
    ]

    for f in flows:
        col = f["col"]
        matched = f["count"] == f["expected"]
        partial = f["count"] > 0 and not matched

        bg  = PANEL_MATCH if matched else (PANEL_WARN if partial else PANEL_MISS)
        txt = GREEN_OK    if matched else (AMBER_TEXT  if partial else RED_TEXT)

        # Label row (row 2)
        cell2 = ws[col + "2"]
        cell2.fill   = fill(bg)
        cell2.font   = Font(bold=True, color=WHITE, size=8, name="Calibri")
        cell2.alignment = center()

        ratio   = "{}/{}".format(f["count"], f["expected"])
        mark    = "✓" if matched else ("!" if partial else "✗")
        cell2.value = "{} {}\n{}  {}".format(f["label"], f["name"], ratio, mark)
        cell2.alignment = Alignment(horizontal="center", vertical="center",
                                     wrap_text=True)

        border_color = "1DB954" if matched else ("FFA726" if partial else "EF5350")
        side = Side(style="medium", color=border_color)
        cell2.border = Border(left=side, right=side, top=side, bottom=side)

    # "Updated" column (L2)
    ws["L2"] = "Synced\n08:00 ✓"
    ws["L2"].fill  = fill(PANEL_BG)
    ws["L2"].font  = Font(bold=False, color=GREY_TEXT, size=8, name="Calibri")
    ws["L2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 32

    # ── Row 3: Column headers ─────────────────────────────────────────────
    headers = ["ID", "Time", "Flow", "Details", "Status", "Error", ""]
    for i, h in enumerate(headers, 1):
        c = ws.cell(3, i)
        c.value = h
        c.fill  = fill(LIGHT_NAVY)
        c.font  = Font(bold=True, color=WHITE, size=9, name="Calibri")
        c.alignment = center()

    # freeze rows 1-3
    ws.freeze_panes = "A4"

    # ── Sample data rows ─────────────────────────────────────────────────
    now = datetime.today()

    def sample_row(row, exec_id, dt, flow, details, status, error="", bg=None, sub=False):
        vals = [
            "" if sub else exec_id,
            dt.strftime("%Y-%m-%d %H:%M") if not sub else "",
            "" if sub else flow,
            details,
            status,
            error
        ]
        bg_map = {
            "F1": F1_BG, "F2": F2_BG, "F3": F3_BG, "F5": F5_BG,
        }
        row_bg = bg or (bg_map.get(flow, WHITE) if not sub else None)

        for i, v in enumerate(vals, 1):
            c = ws.cell(row, i)
            c.value = v
            if sub:
                c.font = Font(color="444444", size=9, name="Calibri")
                c.alignment = left()
                if i == 4:
                    # indent sub items
                    c.alignment = Alignment(indent=2, vertical="center")
                # sub-row status styling
                if i == 5 and status == "Synced":
                    c.fill = fill("D4EDDA")
                elif i == 5 and status == "Failed":
                    c.fill = fill("F8D7DA")
                elif i == 4:
                    c.fill = fill("E8F5E9") if status == "Synced" else fill("FFF8E1")
            else:
                if row_bg:
                    c.fill = fill(row_bg)
                c.font = Font(bold=(not sub), color="1A1A1A", size=9, name="Calibri")
                c.alignment = left()
                if i == 5:
                    c.alignment = center()

    # F1 — PO received
    t1 = now.replace(hour=8, minute=12)
    sample_row(4,  "WK-601", t1, "F1 Receiving",
               "SO-1847  PO received → 3 items WASP",
               "Synced")
    sample_row(5,  "", t1, "", "└ CAR-2OZ  x 24 pcs  RECEIVING-DOCK  add", "Synced", sub=True)
    sample_row(6,  "", t1, "", "└ TGR-4OZ  x 12 pcs  RECEIVING-DOCK  add", "Synced", sub=True)
    sample_row(7,  "", t1, "", "└ LCP-2    x  6 each  RECEIVING-DOCK  add", "Synced", sub=True)

    # F2 — WASP adjustment
    t2 = now.replace(hour=9, minute=44)
    sample_row(8,  "WK-602", t2, "F2 Adjustments",
               "WASP Adjustment  EO-THY x2  WASP → Katana @ PRODUCTION @ MMH Kelowna",
               "Synced")
    sample_row(9,  "", t2, "", "└ EO-THY  x2 grams  PRODUCTION  add  CAR041  2028-06-30", "Synced", sub=True)

    t3 = now.replace(hour=9, minute=51)
    sample_row(10, "WK-603", t3, "F2 Adjustments",
               "WASP Adjustment  LCP-2 x1  WASP → Katana @ PRODUCTION @ MMH Kelowna",
               "Synced")
    sample_row(11, "", t3, "", "└ LCP-2  x1 each  PRODUCTION  remove  CAR023  2028-12-31", "Synced", sub=True)

    # F3 — Stock transfer
    t4 = now.replace(hour=10, minute=5)
    sample_row(12, "WK-604", t4, "F3 Transfers",
               "ST-0091  MMH Kelowna → Storage Warehouse  2 SKUs",
               "Synced")
    sample_row(13, "", t4, "", "└ CAR-2OZ  x 12 pcs  Storage Warehouse  add", "Synced", sub=True)
    sample_row(14, "", t4, "", "└ LCP-2    x  6 each  Storage Warehouse  add", "Synced", sub=True)

    # F5 — ShipStation order — MISSED (shows the discrepancy)
    t5 = now.replace(hour=11, minute=22)
    sample_row(15, "WK-605", t5, "F5 Fulfillment",
               "#ORD-8841  3 items → SHIPPING-DOCK deduction",
               "Synced")
    sample_row(16, "", t5, "", "└ TGR-4OZ  x 2 pcs  SHOPIFY  remove", "Deducted", sub=True)
    sample_row(17, "", t5, "", "└ EO-THY   x 1 g    SHOPIFY  remove", "Deducted", sub=True)

    # More F5 rows to simulate volume
    for idx in range(18, 26):
        hour_off = 11 + (idx - 18) // 3
        min_off  = ((idx - 18) % 3) * 15
        tx = now.replace(hour=hour_off, minute=min_off)
        ord_num = 8842 + (idx - 18)
        sample_row(idx, "WK-{:03d}".format(606 + idx - 18), tx, "F5 Fulfillment",
                   "#ORD-{}  2 items → deduction".format(ord_num), "Synced")

    # ── Legend note (row after data) ─────────────────────────────────────
    note_row = 27
    ws.cell(note_row, 8).value = "■ ✓ = full match"
    ws.cell(note_row, 8).fill  = fill(PANEL_MATCH)
    ws.cell(note_row, 8).font  = Font(color=WHITE, size=8, bold=True, name="Calibri")
    ws.cell(note_row, 8).alignment = center()

    ws.cell(note_row, 9).value = "■ ! = partial"
    ws.cell(note_row, 9).fill  = fill(PANEL_WARN)
    ws.cell(note_row, 9).font  = Font(color=WHITE, size=8, bold=True, name="Calibri")
    ws.cell(note_row, 9).alignment = center()

    ws.cell(note_row, 10).value = "■ ✗ = miss"
    ws.cell(note_row, 10).fill  = fill(PANEL_MISS)
    ws.cell(note_row, 10).font  = Font(color=WHITE, size=8, bold=True, name="Calibri")
    ws.cell(note_row, 10).alignment = center()

    ws.cell(note_row, 11).value = "Counter updates daily 8AM"
    ws.cell(note_row, 11).font  = Font(color=GREY_TEXT, size=8, italic=True, name="Calibri")
    ws.merge_cells("K{}:L{}".format(note_row, note_row))

    # ── Sheet 2: Coverage History ─────────────────────────────────────────
    ws2 = wb.create_sheet("Coverage History")
    ws2.column_dimensions["A"].width = 13
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 10
    ws2.column_dimensions["D"].width = 10
    ws2.column_dimensions["E"].width = 10
    ws2.column_dimensions["F"].width = 10
    ws2.column_dimensions["G"].width = 10
    ws2.column_dimensions["H"].width = 10
    ws2.column_dimensions["I"].width = 10
    ws2.column_dimensions["J"].width = 18

    # Header
    ws2.merge_cells("A1:J1")
    ws2["A1"] = "FLOW COVERAGE HISTORY"
    ws2["A1"].fill = fill(DARK_NAVY)
    ws2["A1"].font = Font(bold=True, color=WHITE, size=12, name="Calibri")
    ws2["A1"].alignment = center()
    ws2.row_dimensions[1].height = 22

    hdrs2 = ["Date",
             "F1 Logged", "F1 Expected",
             "F2 Logged", "F2 Expected",
             "F3 Logged", "F3 Expected",
             "F5 Logged", "F5 Expected",
             "Gaps?"]
    for i, h in enumerate(hdrs2, 1):
        c = ws2.cell(2, i)
        c.value = h
        c.fill  = fill(MID_NAVY)
        c.font  = Font(bold=True, color=WHITE, size=9, name="Calibri")
        c.alignment = center()
    ws2.row_dimensions[2].height = 15
    ws2.freeze_panes = "A3"

    # Sample history rows (last 7 days)
    import random
    random.seed(42)
    history = [
        # date,    F1l, F1e,  F2l, F2e,  F3l, F3e,  F5l,  F5e
        ("2026-02-27",  2,   2,   8,   8,   1,   1,  52,  52),
        ("2026-02-28",  0,   0,   5,   5,   0,   0,  61,  61),
        ("2026-03-01",  0,   0,   0,   0,   0,   0,   0,   0),  # weekend
        ("2026-03-02",  1,   1,   3,   3,   1,   1,  55,  55),
        ("2026-03-03",  3,   3,  11,  11,   2,   2,  48,  48),
        ("2026-03-04",  2,   2,   7,   8,   0,   0,  43,  43),  # F2 miss!
        ("2026-03-05",  3,   3,  12,  12,   2,   2,  45,  47),  # F5 partial
    ]

    for ri, row_data in enumerate(history, 3):
        date_str, f1l, f1e, f2l, f2e, f3l, f3e, f5l, f5e = row_data
        gaps = []
        if f1l != f1e: gaps.append("F1:{}/{}".format(f1l, f1e))
        if f2l != f2e: gaps.append("F2:{}/{}".format(f2l, f2e))
        if f3l != f3e: gaps.append("F3:{}/{}".format(f3l, f3e))
        if f5l != f5e: gaps.append("F5:{}/{}".format(f5l, f5e))
        gap_text = "  ".join(gaps) if gaps else "✓ All matched"

        vals = [date_str, f1l, f1e, f2l, f2e, f3l, f3e, f5l, f5e, gap_text]
        row_bg = "FFEBEE" if gaps else "E8F5E9"
        gap_fg = RED_TEXT  if gaps else GREEN_OK

        for ci, v in enumerate(vals, 1):
            c = ws2.cell(ri, ci)
            c.value = v
            c.font  = Font(size=9, name="Calibri",
                           bold=(ci == 1),
                           color=gap_fg if ci == 10 else "1A1A1A")
            c.fill  = fill(row_bg)
            c.alignment = center()
            if ci == 10:
                c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        ws2.row_dimensions[ri].height = 14

    # ── Sheet 3: HOW IT WORKS ─────────────────────────────────────────────
    ws3 = wb.create_sheet("How It Works")
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 55
    ws3.column_dimensions["C"].width = 40

    ws3.merge_cells("A1:C1")
    ws3["A1"] = "FLOW COVERAGE — HOW THE COUNT WORKS"
    ws3["A1"].fill = fill(DARK_NAVY)
    ws3["A1"].font = Font(bold=True, color=WHITE, size=12, name="Calibri")
    ws3["A1"].alignment = center()
    ws3.row_dimensions[1].height = 22

    explanation = [
        # [Flow, How we count LOGGED, How we count EXPECTED (source of truth)]
        ["Flow", "Logged  (Activity tab)", "Expected  (Source of truth)"],
        ["F1 – PO Receiving",
         "Count Activity rows with 'F1 Receiving' and today's date",
         "GET /purchase_orders?status=received  → filter updated_at = today"],
        ["F2 – WASP Adjustments",
         "Count Activity rows with 'F2 Adjustments' and today's date",
         "GET /transactions/item/list from WASP API  → count today's records"],
        ["F3 – Stock Transfers",
         "Count Activity rows with 'F3 Transfers' and today's date",
         "GET /stock_transfers?status=completed  → filter updated_at = today"],
        ["F5 – Fulfillment",
         "Count Activity rows with 'F5 Fulfillment' and today's date",
         "GET /shipments?shipDateStart=today from ShipStation API"],
        ["", "", ""],
        ["⚡ Update schedule",
         "Daily time trigger runs countFlowCoverage() at 08:00\nWrites counts to Activity sheet H1:L2",
         "Alert fires to #it-support Slack if any flow has gaps"],
        ["", "", ""],
        ["⚡ Reliability gaps",
         "Katana webhooks have NO retry — if GAS cold-starts, event is lost",
         "Counter detects the miss.  Recovery: manual re-poll or re-trigger from Katana"],
        ["⚡ Reducing misses",
         "1.  Warmup trigger (ping every 5 min) keeps GAS warm\n"
         "2.  F3 + F5 use polling (not webhook) — inherently reliable\n"
         "3.  F1/F4 use webhooks — most at risk",
         "Future: store Katana event IDs and cross-check nightly"],
    ]

    for ri, row_data in enumerate(explanation, 2):
        is_header = ri == 2
        for ci, v in enumerate(row_data, 1):
            c = ws3.cell(ri, ci)
            c.value = v
            c.alignment = Alignment(horizontal="left", vertical="top",
                                    wrap_text=True)
            if is_header:
                c.fill = fill(MID_NAVY)
                c.font = Font(bold=True, color=WHITE, size=9, name="Calibri")
            elif v.startswith("⚡"):
                c.fill = fill("FFF3E0")
                c.font = Font(bold=True, color="E65100", size=9, name="Calibri")
            else:
                c.fill = fill("F5F5F5" if ri % 2 == 0 else WHITE)
                c.font = Font(size=9, name="Calibri", color="1A1A1A")
        ws3.row_dimensions[ri].height = 40 if ri > 6 else 20

    ws3.freeze_panes = "A3"

    wb.save(path)
    print("Saved:", path)


if __name__ == "__main__":
    out = r"C:\Users\Admin\Documents\claude-projects\wasp-katana\docs\flow-coverage-demo.xlsx"
    build(out)
