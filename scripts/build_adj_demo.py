"""
Adj Demo — modifies only the Match column (O) in the Wasp tab.

Match values after this script:
  MATCH        — row qty == K.Qty (unchanged)
  SPLIT MATCH  — row looks mismatched but site W.Total == K.Qty (already balanced)
  ADJUSTABLE   — truly mismatched, site W.Total != K.Qty, auto-fixable (no lot, no UOM conversion)
  MISMATCH     — truly mismatched but needs manual handling (lot-tracked or UOM conversion)
  WASP ONLY    — exists in WASP but not Katana (unchanged)
  ZERO         — both zero (unchanged)
"""

import openpyxl
from collections import defaultdict

SRC = r"C:\Users\Admin\Documents\claude-projects\wasp-katana\docs\_2026_Katana-Wasp-Inventory-Sync-latest.xlsx"
DST = r"C:\Users\Admin\Documents\claude-projects\wasp-katana\docs\adj-demo.xlsx"

COL_SKU     = 1   # A
COL_SITE    = 4   # D
COL_LOT_TRK = 6   # F
COL_UOM     = 9   # I  — contains "→" when purchase UOM conversion exists
COL_KQTY    = 10  # J
COL_WQTY    = 12  # L
COL_MATCH   = 15  # O
COL_NET_ADJ = 19  # S  — after existing hidden cols (Orig Lot=17, Orig DateCode=18)


def safe_float(v):
    try:
        return float(v) if v not in (None, "") else 0.0
    except (ValueError, TypeError):
        return 0.0


def is_sep(row_vals):
    return " > " in str(row_vals[COL_SKU - 1] or "")


def main():
    wb = openpyxl.load_workbook(SRC)
    ws = wb["Wasp"]

    # Pass 1 — build site W.Total per SKU|site
    rows_vals = [list(r) for r in ws.iter_rows(min_row=2, values_only=True)]

    site_totals = defaultdict(float)
    for rv in rows_vals:
        if is_sep(rv):
            continue
        sku  = str(rv[COL_SKU  - 1] or "").strip()
        site = str(rv[COL_SITE - 1] or "").strip()
        wqty = safe_float(rv[COL_WQTY - 1])
        if sku and site:
            site_totals[sku + "|" + site] += wqty

    # Write Net Adj header
    from openpyxl.styles import PatternFill, Font, Alignment
    hdr = ws.cell(row=1, column=COL_NET_ADJ, value="Net Adj")
    hdr.fill = PatternFill("solid", fgColor="263238")
    hdr.font = Font(color="FFFFFF", bold=True, size=10)
    hdr.alignment = Alignment(horizontal="center")
    ws.column_dimensions["S"].width = 10

    # Pass 2 — first determine new_match for every row, record which rows are ADJUSTABLE per group
    # group_rows[sku|site] = [row_index, ...]  (only ADJUSTABLE rows)
    group_rows   = defaultdict(list)
    row_matches  = {}   # ri -> new_match
    row_adj_dir  = {}   # ri -> "+" or "-"

    for ri, rv in enumerate(rows_vals, start=2):
        if is_sep(rv) or not str(rv[COL_SKU - 1] or "").strip():
            continue

        sku     = str(rv[COL_SKU     - 1] or "").strip()
        site    = str(rv[COL_SITE    - 1] or "").strip()
        lot_trk = str(rv[COL_LOT_TRK - 1] or "").strip()
        uom     = str(rv[COL_UOM     - 1] or "").strip()
        kqty    = safe_float(rv[COL_KQTY - 1])
        match   = str(rv[COL_MATCH   - 1] or "").strip()

        site_total = round(site_totals.get(sku + "|" + site, 0.0), 4)
        adj_qty    = round(kqty - site_total, 4)

        if match == "MISMATCH":
            if abs(adj_qty) < 0.01:
                new_match = "SPLIT MATCH"
            elif lot_trk == "Yes" or "→" in uom:
                new_match = "MISMATCH"
            else:
                new_match = "ADJUSTABLE"
        else:
            new_match = match

        row_matches[ri] = new_match

        if new_match == "ADJUSTABLE":
            row_adj_dir[ri] = "+" if adj_qty > 0 else "-"
            group_rows[sku + "|" + site].append(ri)

    # Build next-row map: each row points to the next in its group (circular)
    next_row = {}
    for key, rows in group_rows.items():
        if len(rows) > 1:
            for i, ri in enumerate(rows):
                next_row[ri] = rows[(i + 1) % len(rows)]

    # Pass 3 — write to sheet
    counts = defaultdict(int)

    for ri, rv in enumerate(rows_vals, start=2):
        if is_sep(rv) or not str(rv[COL_SKU - 1] or "").strip():
            continue

        new_match = row_matches.get(ri)
        if new_match is None:
            continue

        # Update Match cell
        if str(rv[COL_MATCH - 1] or "").strip() != new_match:
            ws.cell(row=ri, column=COL_MATCH).value = new_match
        counts[new_match] += 1

        # Net Adj: "+" or "-" with hyperlink to next connected row
        if new_match == "ADJUSTABLE":
            cell = ws.cell(row=ri, column=COL_NET_ADJ, value=row_adj_dir[ri])
            if ri in next_row:
                target = next_row[ri]
                cell.hyperlink = "#Wasp!A" + str(target)
                cell.style = "Hyperlink"

    wb.save(DST)
    print("Saved:", DST)
    for status in ("MATCH", "SPLIT MATCH", "ADJUSTABLE", "MISMATCH", "WASP ONLY", "ZERO"):
        if counts[status]:
            print(f"  {status}: {counts[status]}")


if __name__ == "__main__":
    main()
