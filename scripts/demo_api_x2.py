"""
Demo XLSX — API Key x2 proposal for WASP/Katana rows
Shows BEFORE / AFTER comparison for the Health Tracker
"""
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "API Key x2 Demo"

# ── Colours ────────────────────────────────────────────────────────────────
DARK_HEADER  = "263238"
WHITE        = "FFFFFF"
LIGHT_ROW    = "F5F7F9"
SEP_SHEET    = "1B5E20"
CHIP_WASP_BG = "C8E6C9"
CHIP_WASP_FG = "1B5E20"
YELLOW_WARN  = "553600"
YELLOW_TEXT  = "FBBF24"
RED_WARN     = "5C0011"
RED_TEXT     = "F87171"
BLUE_LINK    = "1155CC"
BEFORE_BG    = "FAFAFA"
AFTER_BG     = "F0FAF0"
LABEL_BEFORE = "BDBDBD"   # grey label
LABEL_AFTER  = "388E3C"   # green label

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(color=None, bold=False, italic=False, size=10):
    return Font(color=color or "000000", bold=bold, italic=italic, size=size)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── Column layout (subset of the 19 cols — focused on auth/expiry) ─────────
# We show: System Name | Auth Type | Auth Location | Expiry Date | Days Left | Notes
COLS = [
    ("System Name",    28),
    ("Auth Type",      14),
    ("Auth Location",  26),
    ("Expiry Date",    14),
    ("Days Left",      10),
    ("Notes",          46),
]

for ci, (header, width) in enumerate(COLS, start=1):
    ws.column_dimensions[get_column_letter(ci)].width = width

ws.row_dimensions[1].height = 14
ws.row_dimensions[2].height = 28
ws.row_dimensions[3].height = 20
ws.row_dimensions[4].height = 36
ws.row_dimensions[5].height = 20
ws.row_dimensions[6].height = 36
ws.row_dimensions[7].height = 20
ws.row_dimensions[8].height = 36

# ── Row 1 — Title ───────────────────────────────────────────────────────────
ws.merge_cells("A1:F1")
ws["A1"] = "Health Tracker — API Key x2 Demo  (WASP/Katana rows)"
ws["A1"].fill    = fill(DARK_HEADER)
ws["A1"].font    = font(WHITE, bold=True, size=11)
ws["A1"].alignment = left()

# ── Row 2 — Column headers ──────────────────────────────────────────────────
for ci, (header, _) in enumerate(COLS, start=1):
    cell = ws.cell(row=2, column=ci, value=header)
    cell.fill      = fill(DARK_HEADER)
    cell.font      = font(WHITE, bold=True, size=10)
    cell.alignment = center()
    cell.border    = border

# ── Helper: write a styled data row ─────────────────────────────────────────
def write_row(row_num, values, row_bg, label=None, label_color=None):
    for ci, val in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=ci, value=val)
        cell.fill      = fill(row_bg)
        cell.font      = font(size=10)
        cell.alignment = left()
        cell.border    = border
    # Auth Type chip (col 2)
    auth_cell = ws.cell(row=row_num, column=2)
    if "x2" in str(auth_cell.value):
        auth_cell.fill = fill("FFE0B2")
        auth_cell.font = font("BF360C", bold=True, size=10)
        auth_cell.alignment = center()
    # Days Left colouring (col 5)
    dl_cell = ws.cell(row=row_num, column=5)
    if isinstance(dl_cell.value, int):
        if dl_cell.value < 0:
            dl_cell.fill = fill(RED_WARN)
            dl_cell.font = font(RED_TEXT, bold=True)
        elif dl_cell.value <= 30:
            dl_cell.fill = fill(YELLOW_WARN)
            dl_cell.font = font(YELLOW_TEXT, bold=True)
        dl_cell.alignment = center()
    elif dl_cell.value in ("—", "NO EXPIRY"):
        dl_cell.font = font("546E7A", bold=False)
        dl_cell.alignment = center()
    # Label in last col if provided
    if label:
        label_cell = ws.cell(row=row_num, column=6)
        label_cell.font = font(label_color or "000000", bold=True, size=9)

# ── SEPARATOR — GOOGLE SHEET ─────────────────────────────────────────────────
ws.merge_cells("A3:F3")
ws["A3"] = "  GOOGLE SHEET"
ws["A3"].fill      = fill(SEP_SHEET)
ws["A3"].font      = font(WHITE, bold=True, size=9)
ws["A3"].alignment = left()

# ── Row 4 — BEFORE label row (greyed out) ────────────────────────────────────
ws.merge_cells("A4:F4")  # actually don't merge — just label col A
ws.unmerge_cells("A4:F4")
label4 = ws.cell(row=4, column=1, value="◀  BEFORE")
label4.fill = fill("EEEEEE")
label4.font = font(LABEL_BEFORE, bold=True, size=9)
label4.alignment = center()
for ci in range(2, 7):
    ws.cell(row=4, column=ci).fill = fill("EEEEEE")
    ws.cell(row=4, column=ci).border = border

# ── Row 5 — BEFORE: current state ───────────────────────────────────────────
write_row(5, [
    "2026_Katana-Wasp Inventory Sync",
    "API Key",
    "Script Properties",
    "",          # blank expiry
    "",          # blank days left
    "",          # no notes
], BEFORE_BG)
# grey out the blank cells to show the gap
for ci in [4, 5, 6]:
    c = ws.cell(row=5, column=ci)
    c.fill = fill("EEEEEE")
    c.font = font("9E9E9E", italic=True)
    c.value = "(empty)" if ci < 6 else ""

# ── Row 6 — AFTER label ──────────────────────────────────────────────────────
label6 = ws.cell(row=6, column=1, value="▶  AFTER  (proposed)")
label6.fill = fill("E8F5E9")
label6.font = font(LABEL_AFTER, bold=True, size=9)
label6.alignment = center()
for ci in range(2, 7):
    ws.cell(row=6, column=ci).fill = fill("E8F5E9")
    ws.cell(row=6, column=ci).border = border

# ── Row 7 — AFTER: proposed state ───────────────────────────────────────────
write_row(7, [
    "2026_Katana-Wasp Inventory Sync",
    "API Key x2",
    "Script Properties",
    "2027-12-31",
    302,
    "Katana API: No expiry  |  WASP API: expires 2027-12-31",
], AFTER_BG)

# Days Left auto-note
dl = ws.cell(row=7, column=5)
dl.value = 302
dl.alignment = center()

# Notes cell — style key parts
notes = ws.cell(row=7, column=6)
notes.font = font("546E7A", size=9)

# ── Row 8 — explanation ──────────────────────────────────────────────────────
ws.merge_cells("A8:F8")
exp = ws["A8"]
exp.value = (
    "Days Left (col K) = Expiry Date − TODAY()  →  recalculates automatically each day.  "
    "Turns yellow at ≤30 days, red when expired.  "
    "Auth Type 'API Key x2' signals two separate keys.  Notes column holds both details."
)
exp.fill      = fill("E3F2FD")
exp.font      = font("0D47A1", size=9, italic=True)
exp.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[8].height = 32

out_path = r"C:\Users\Admin\Desktop\api_key_x2_demo.xlsx"
wb.save(out_path)
print("Saved:", out_path)
