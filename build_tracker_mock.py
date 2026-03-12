"""
build_tracker_mock.py
Generates tracker-mock.xlsx modelling the Google Sheets tracker
used by the Katana MRP <-> WASP InventoryCloud sync system.
"""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# Shared colour palette
# Activity tab uses ACTIVITY_FLOW_BG (matches real Google Sheet).
# Flow detail tabs (F1-F5) keep original FLOW_BG (unchanged).
# ─────────────────────────────────────────────

# Original palette — kept for F1-F5 detail tabs
FLOW_BG = {
    'F1': 'CCE5FF',   # blue   – Receiving
    'F2': 'B2DFDB',   # teal   – Adjustments
    'F3': 'FCE4D6',   # peach  – Transfers
    'F4': 'F0E6F6',   # purple – Manufacturing
    'F5': 'D1ECF1',   # cyan   – Shipping
}

# Activity-tab row background colours (matches target Google Sheet screenshot)
ACTIVITY_FLOW_BG = {
    'F1': 'BBDEFB',   # light blue      – Receiving
    'F2': 'B2DFDB',   # light teal      – Adjustments
    'F3': 'FFE0B2',   # light peach/orange – Transfers
    'F4': 'FFF9C4',   # light yellow    – Manufacturing
    'F5': 'E8D5F5',   # light lavender  – Shipping
}

# Status cell fills — Activity tab (per-status palette from spec)
ACTIVITY_STATUS_FILL = {
    'Complete':  PatternFill('solid', fgColor='C8E6C9'),   # light green
    'Received':  PatternFill('solid', fgColor='A5D6A7'),   # green
    'Synced':    PatternFill('solid', fgColor='A5D6A7'),   # green
    'Added':     PatternFill('solid', fgColor='C8E6C9'),   # light green
    'Produced':  PatternFill('solid', fgColor='B2DFDB'),   # teal
    'Consumed':  PatternFill('solid', fgColor='C8E6C9'),   # light green
    'Picked':    PatternFill('solid', fgColor='C8E6C9'),   # light green
    'Shipped':   PatternFill('solid', fgColor='A5D6A7'),   # green
    'Failed':    PatternFill('solid', fgColor='FFCDD2'),   # light red/pink
    'Partial':   PatternFill('solid', fgColor='FFE0B2'),   # orange
    'Skipped':   PatternFill('solid', fgColor='FFF9C4'),   # amber/yellow
    # lowercase variants used on parent rows
    'success':   PatternFill('solid', fgColor='C8E6C9'),
    'failed':    PatternFill('solid', fgColor='FFCDD2'),
    'partial':   PatternFill('solid', fgColor='FFE0B2'),
    'skipped':   PatternFill('solid', fgColor='FFF9C4'),
}

# Status fill used by flow detail tabs (F1-F5) — unchanged
STATUS_FILL = {
    'success':  PatternFill('solid', fgColor='D4EDDA'),
    'failed':   PatternFill('solid', fgColor='F8D7DA'),
    'partial':  PatternFill('solid', fgColor='FFF3CD'),
    'skipped':  PatternFill('solid', fgColor='FFF8E1'),
    'Received': PatternFill('solid', fgColor='D4EDDA'),
    'Synced':   PatternFill('solid', fgColor='D4EDDA'),
    'Complete': PatternFill('solid', fgColor='D4EDDA'),
    'Shipped':  PatternFill('solid', fgColor='D4EDDA'),
    'Failed':   PatternFill('solid', fgColor='F8D7DA'),
    'Partial':  PatternFill('solid', fgColor='FFF3CD'),
    'Skipped':  PatternFill('solid', fgColor='FFF8E1'),
    'Added':    PatternFill('solid', fgColor='D4EDDA'),
    'Picked':   PatternFill('solid', fgColor='D4EDDA'),
    'PENDING':  PatternFill('solid', fgColor='FFF8E1'),
    'Complete_adj': PatternFill('solid', fgColor='D4EDDA'),
    'COMPLETE': PatternFill('solid', fgColor='D4EDDA'),
}

SUB_OK_FILL   = PatternFill('solid', fgColor='E8F5E9')   # light green
SUB_ERR_FILL  = PatternFill('solid', fgColor='FFEBEE')   # light red
SUB_SKIP_FILL = PatternFill('solid', fgColor='FFF8E1')   # light amber
ERR_CELL_FILL = PatternFill('solid', fgColor='FFF0F0')   # very light red
HEADER_COL_FILL = PatternFill('solid', fgColor='E8EAF6') # indigo-50 (flow tabs)
DARK_HEADER_FILL = PatternFill('solid', fgColor='263238') # dark blue-grey (Activity rows 1+3)

BOLD = Font(bold=True)
GREY_FONT = Font(color='666666')

def _fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def _status_fill(status_str):
    return STATUS_FILL.get(status_str, PatternFill('solid', fgColor='FFFFFF'))

def freeze_and_autosize(ws, freeze_row=1, col_widths=None):
    """Freeze top N rows and set column widths."""
    ws.freeze_panes = ws.cell(row=freeze_row + 1, column=1)
    if col_widths:
        for col_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width


# ═══════════════════════════════════════════════════════════════
# TAB 1 – Activity
# Columns: ID | Time | Flow | Details | Status | Error | Retry
# Rows 1-3: CC header (frozen), data from row 4
# ═══════════════════════════════════════════════════════════════

def build_activity(wb):
    ws = wb.create_sheet('Activity')

    # ── CC header rows 1-3 ──────────────────────────────────────
    ws.append(['ACTIVITY LOG', '', '', 'Katana-WASP Inventory Sync', '', '', ''])
    ws['A1'].font = Font(bold=True, size=12, color='FFFFFF')
    ws.append(['All Flows — Chronological', '', '', '', '', '', ''])
    ws['A2'].font = Font(color='FFFFFF')
    ws.append(['ID', 'Time', 'Flow', 'Details', 'Status', 'Error', 'Retry'])
    # Apply dark background to all 3 header rows
    for row_num in [1, 2, 3]:
        for cell in ws[row_num]:
            cell.fill = DARK_HEADER_FILL
    # Header row 3: bold white text
    for cell in ws[3]:
        cell.font = Font(bold=True, color='FFFFFF')

    freeze_and_autosize(ws, freeze_row=3,
        col_widths=[10, 16, 18, 64, 10, 36, 6])

    # Helper: write a header row + optional sub-items
    def write_event(exec_id, ts, flow, flow_label, details, status_text, error='', sub_items=None):
        """
        sub_items: list of dicts with keys:
            sku, qty, action, status, error, is_last, nested, is_parent, batch_count
        """
        fill = _fill(ACTIVITY_FLOW_BG[flow])
        row = [exec_id, ts, flow_label, details, status_text, error, '']
        ws.append(row)
        r = ws.max_row
        for c in range(1, 8):
            ws.cell(r, c).fill = fill
        ws.cell(r, 1).font = Font(bold=True)
        # Status cell — use Activity-tab status palette
        act_status_fill = ACTIVITY_STATUS_FILL.get(status_text)
        if act_status_fill:
            ws.cell(r, 5).fill = act_status_fill
        if error:
            ws.cell(r, 6).fill = ERR_CELL_FILL

        if sub_items:
            for idx, si in enumerate(sub_items):
                is_last = (idx == len(sub_items) - 1)
                tree = '└─' if is_last else '├─'

                if si.get('nested'):
                    nxt_nested = (idx + 1 < len(sub_items)) and sub_items[idx + 1].get('nested')
                    nl = '├─' if nxt_nested else '└─'
                    sub_detail = f'    │   {nl} x{si["qty"]}'
                    if si.get('action'):
                        sub_detail += f'  {si["action"]}'
                elif si.get('is_parent'):
                    sub_detail = f'    {tree} {si["sku"]} x{si["qty"]}'
                    if si.get('action'):
                        sub_detail += f'  {si["action"]}'
                    if si.get('batch_count'):
                        sub_detail += f' ({si["batch_count"]} batches)'
                else:
                    sub_detail = f'    {tree} {si["sku"]} x{si["qty"]}'
                    if si.get('action'):
                        sub_detail += f'  {si["action"]}'

                sub_status = si.get('status', '')
                sub_err = si.get('error', '')
                ws.append(['', '', '', sub_detail, sub_status, sub_err, ''])
                sr = ws.max_row
                # Detail cell fill
                if sub_status == 'Skipped':
                    ws.cell(sr, 4).fill = SUB_SKIP_FILL
                elif sub_err:
                    ws.cell(sr, 4).fill = SUB_ERR_FILL
                else:
                    ws.cell(sr, 4).fill = SUB_OK_FILL
                # Status cell fill — use Activity-tab status palette
                act_sub_fill = ACTIVITY_STATUS_FILL.get(sub_status)
                if act_sub_fill:
                    ws.cell(sr, 5).fill = act_sub_fill
                if sub_err:
                    ws.cell(sr, 6).fill = ERR_CELL_FILL

    # ── DATA ROWS ───────────────────────────────────────────────
    # F1 – PO Received, 3 items, success
    write_event('WK-001', '2026-02-24 08:03', 'F1', 'F1 Receiving',
        'PO-1042  3 items  → RECEIVING-DOCK, PRODUCTION @ MMH Kelowna',
        'Received', sub_items=[
            {'sku': 'UFC-4OZ',     'qty': 200, 'action': 'RECEIVING-DOCK  lot:L2025-041', 'status': 'Received'},
            {'sku': 'B-WAX-500G',  'qty': 50,  'action': 'PRODUCTION  lot:L2025-039',    'status': 'Received'},
            {'sku': 'COLLAGEN-30', 'qty': 100, 'action': 'PRODUCTION  lot:L2025-040',    'status': 'Received'},
        ])

    # F2 – WASP→Katana single item add, success
    write_event('WK-002', '2026-02-24 09:17', 'F2', 'F2 Adjustments',
        'UFC-8OZ x48  lot:L2025-112  WASP → Katana @ RECEIVING-DOCK @ MMH Kelowna',
        'Synced')

    # F4 – MO Complete, multi-ingredient remove + FG add
    write_event('WK-003', '2026-02-24 10:45', 'F4', 'F4 Manufacturing',
        'MO-7246  UFC-4OZ x240  → PROD-RECEIVING @ MMH Kelowna',
        'Complete', sub_items=[
            {'sku': 'B-WAX-500G',  'qty': 12,  'action': 'PRODUCTION  lot:L2025-039  remove', 'status': 'Added'},
            {'sku': 'COLLAGEN-30', 'qty': 8,   'action': 'PRODUCTION  lot:L2025-040  remove', 'status': 'Added'},
            {'sku': 'UFC-4OZ',     'qty': 240, 'action': '→ PROD-RECEIVING  lot:UFC410B',      'status': 'Added'},
        ])

    # F5 – SO Delivered / Shopify fulfillment, success
    write_event('WK-004', '2026-02-24 11:02', 'F5', 'F5 Shipping',
        '#90412  2 items  → SHIPPING-DOCK',
        'Shipped', sub_items=[
            {'sku': 'UFC-4OZ', 'qty': 2, 'action': 'SHIPPING-DOCK', 'status': 'Picked'},
            {'sku': 'UFC-8OZ', 'qty': 1, 'action': 'SHIPPING-DOCK', 'status': 'Picked'},
        ])

    # F2 – multi-item batch remove, partial failure
    write_event('WK-005', '2026-02-24 12:30', 'F2', 'F2 Adjustments',
        '3 items  WASP → Katana @ PRODUCTION @ MMH Kelowna',
        'Partial', error='1 failed', sub_items=[
            {'sku': 'B-WAX-500G',  'qty': 5, 'action': 'PRODUCTION  lot:L2025-039  remove', 'status': 'Synced'},
            {'sku': 'COLLAGEN-30', 'qty': 3, 'action': 'PRODUCTION  lot:L2025-040  remove', 'status': 'Synced'},
            {'sku': 'OP-195',      'qty': 2, 'action': 'PRODUCTION  remove',                'status': 'Failed',
             'error': 'Item not in Katana'},
        ])

    # F1 – PO Received, multi-batch single SKU + second SKU
    write_event('WK-006', '2026-02-24 13:05', 'F1', 'F1 Receiving',
        'PO-1047  2 items  → RECEIVING-DOCK @ MMH Kelowna',
        'Received', sub_items=[
            {'sku': 'UFC-8OZ', 'qty': 500, 'action': 'RECEIVING-DOCK',
             'is_parent': True, 'batch_count': 2, 'status': ''},
            {'sku': '',   'qty': 300, 'action': 'lot:L2025-111  exp:2027-11-30', 'status': 'Received', 'nested': True},
            {'sku': '',   'qty': 200, 'action': 'lot:L2025-112  exp:2027-12-31', 'status': 'Received', 'nested': True},
            {'sku': 'B-WAX-500G', 'qty': 40, 'action': 'PRODUCTION  lot:L2025-039', 'status': 'Received'},
        ])

    # F3 – Amazon Transfer / SO Delivered
    write_event('WK-007', '2026-02-24 14:20', 'F3', 'F3 Transfers',
        'SO-88231  B-WAX-500G x20  → SHOPIFY',
        'Complete', sub_items=[
            {'sku': 'B-WAX-500G', 'qty': 20, 'action': '→ SHOPIFY', 'status': 'Complete'},
        ])

    # F4 – MO failed (WASP error)
    write_event('WK-008', '2026-02-24 15:10', 'F4', 'F4 Manufacturing',
        'MO-7251  COLLAGEN-30 x120  Failed',
        'Failed', error='Insufficient qty at location', sub_items=[
            {'sku': 'B-WAX-500G',  'qty': 6, 'action': 'PRODUCTION  remove', 'status': 'Failed',
             'error': 'Insufficient qty at location'},
            {'sku': 'COLLAGEN-30', 'qty': 120, 'action': '→ PROD-RECEIVING', 'status': 'Failed',
             'error': 'Insufficient qty at location'},
        ])

    # F2 – single add, skipped (recently synced)
    write_event('WK-009', '2026-02-24 16:02', 'F2', 'F2 Adjustments',
        'UFC-4OZ x10  WASP → Katana @ SW-STORAGE @ Storage Warehouse',
        'Skipped')

    # F5 – large Shopify order, 4 items
    write_event('WK-010', '2026-02-24 17:45', 'F5', 'F5 Shipping',
        '#90589  4 items  → SHIPPING-DOCK',
        'Shipped', sub_items=[
            {'sku': 'UFC-4OZ',    'qty': 6,  'action': 'SHIPPING-DOCK', 'status': 'Picked'},
            {'sku': 'UFC-8OZ',    'qty': 3,  'action': 'SHIPPING-DOCK', 'status': 'Picked'},
            {'sku': 'B-WAX-500G', 'qty': 2,  'action': 'SHIPPING-DOCK', 'status': 'Picked'},
            {'sku': 'COLLAGEN-30','qty': 1,  'action': 'SHIPPING-DOCK', 'status': 'Picked'},
        ])

    # F1 – PO Received, 3 items, 1 failed (OP- prefix skipped), 1 not in WASP
    write_event('WK-011', '2026-02-25 08:55', 'F1', 'F1 Receiving',
        'PO-1049  3 items  → RECEIVING-DOCK @ MMH Kelowna  1 failed',
        'Partial', error='1 item not in WASP', sub_items=[
            {'sku': 'UFC-4OZ',  'qty': 30,  'action': 'RECEIVING-DOCK  lot:L2025-041', 'status': 'Received'},
            {'sku': 'OP-195',   'qty': 50,  'action': 'RECEIVING-DOCK',                'status': 'Skipped',
             'error': 'Skipped prefix'},
            {'sku': 'UFC-8OZ',  'qty': 20,  'action': 'RECEIVING-DOCK',                'status': 'Failed',
             'error': 'Item not in WASP'},
        ])

    # F2 – multi-item add from Storage Warehouse
    write_event('WK-012', '2026-02-25 09:30', 'F2', 'F2 Adjustments',
        '2 items  WASP → Katana @ SW-STORAGE @ Storage Warehouse',
        'Synced', sub_items=[
            {'sku': 'UFC-4OZ', 'qty': 24, 'action': 'SW-STORAGE  lot:L2025-041  add', 'status': 'Synced'},
            {'sku': 'UFC-8OZ', 'qty': 12, 'action': 'SW-STORAGE  lot:L2025-112  add', 'status': 'Synced'},
        ])

    # F4 – MO Complete, 2 ingredients + FG, partial (one ingredient failed)
    write_event('WK-013', '2026-02-25 10:15', 'F4', 'F4 Manufacturing',
        'MO-7260  UFC-8OZ x180  → PROD-RECEIVING @ MMH Kelowna',
        'Partial', error='1 failed', sub_items=[
            {'sku': 'B-WAX-500G',  'qty': 9,  'action': 'PRODUCTION  lot:L2025-039  remove', 'status': 'Added'},
            {'sku': 'COLLAGEN-30', 'qty': 6,  'action': 'PRODUCTION  remove',                'status': 'Failed',
             'error': 'Lot not found at location'},
            {'sku': 'UFC-8OZ',     'qty': 180,'action': '→ PROD-RECEIVING  lot:UFC412A',      'status': 'Added'},
        ])

    # F3 – Amazon transfer, 2 items
    write_event('WK-014', '2026-02-25 11:40', 'F3', 'F3 Transfers',
        'SO-88290  2 items  → SHOPIFY',
        'Complete', sub_items=[
            {'sku': 'UFC-4OZ',    'qty': 15, 'action': '→ SHOPIFY', 'status': 'Complete'},
            {'sku': 'COLLAGEN-30','qty': 10, 'action': '→ SHOPIFY', 'status': 'Complete'},
        ])

    # F5 – single item, failed (WASP pick order error)
    write_event('WK-015', '2026-02-25 14:22', 'F5', 'F5 Shipping',
        '#90601  UFC-4OZ x3  Duplicate order number',
        'Failed', error='Duplicate order number')

    # F2 – site not mapped error
    write_event('WK-016', '2026-02-25 15:05', 'F2', 'F2 Adjustments',
        'UFC-4OZ x5  Site not mapped  WASP → Katana',
        'Failed', error='Site not mapped')

    # F1 – PO Received, 3 items, success
    write_event('WK-017', '2026-02-26 08:10', 'F1', 'F1 Receiving',
        'PO-1053  3 items  → PRODUCTION, RECEIVING-DOCK @ MMH Kelowna',
        'Received', sub_items=[
            {'sku': 'COLLAGEN-30', 'qty': 200, 'action': 'PRODUCTION  lot:LOT-20260201',    'status': 'Received'},
            {'sku': 'B-WAX-500G',  'qty': 80,  'action': 'PRODUCTION  lot:L2026-003',        'status': 'Received'},
            {'sku': 'UFC-4OZ',     'qty': 60,  'action': 'RECEIVING-DOCK  lot:L2026-001',    'status': 'Received'},
        ])

    # F4 – MO complete, 3 ingredients, full success
    write_event('WK-018', '2026-02-26 09:45', 'F4', 'F4 Manufacturing',
        'MO-7271  UFC-4OZ x360  → PROD-RECEIVING @ MMH Kelowna',
        'Complete', sub_items=[
            {'sku': 'B-WAX-500G',  'qty': 18, 'action': 'PRODUCTION  lot:L2025-039  remove', 'status': 'Added'},
            {'sku': 'COLLAGEN-30', 'qty': 12, 'action': 'PRODUCTION  lot:L2025-040  remove', 'status': 'Added'},
            {'sku': 'OP-195',      'qty': 2,  'action': 'PRODUCTION  remove',                'status': 'Skipped'},
            {'sku': 'UFC-4OZ',     'qty': 360,'action': '→ PROD-RECEIVING  lot:UFC413C  exp:2029-02-28', 'status': 'Added'},
        ])

    return ws


# ═══════════════════════════════════════════════════════════════
# FLOW DETAIL TABS  F1–F5
# Columns: Exec ID | Time | Ref# | Details | Status | Error
# Rows 1-3: CC header (frozen), data from row 4
# ═══════════════════════════════════════════════════════════════

def _cc_header(ws, flow, title_suffix):
    """Write 3-row CC header for a flow detail tab."""
    flow_name = {
        'F1': 'F1 Receiving',
        'F2': 'F2 Adjustments',
        'F3': 'F3 Transfers',
        'F4': 'F4 Manufacturing',
        'F5': 'F5 Shipping',
    }[flow]
    ws.append([flow_name, '', '', title_suffix, '', ''])
    ws['A1'].font = Font(bold=True, size=12)
    ws.append(['Katana-WASP Sync — ' + flow_name, '', '', '', '', ''])
    ws.append(['Exec ID', 'Time', 'Ref#', 'Details', 'Status', 'Error'])
    for cell in ws[3]:
        cell.font = BOLD
        cell.fill = HEADER_COL_FILL
    freeze_and_autosize(ws, freeze_row=3,
        col_widths=[10, 16, 12, 72, 10, 36])


def _write_flow_event(ws, flow, exec_id, ts, ref, detail, status, error='', sub_items=None):
    """Write a header row + sub-item tree rows to a flow tab."""
    fill = _fill(FLOW_BG[flow])
    ws.append([exec_id, ts, ref, detail, status, error])
    r = ws.max_row
    for c in range(1, 7):
        ws.cell(r, c).fill = fill
    ws.cell(r, 1).font = Font(bold=True)
    ws.cell(r, 5).fill = _status_fill(status)
    if error:
        ws.cell(r, 6).fill = ERR_CELL_FILL

    if sub_items:
        for idx, si in enumerate(sub_items):
            is_last = (idx == len(sub_items) - 1)
            tree = '└─' if is_last else '├─'

            if si.get('nested'):
                nxt = (idx + 1 < len(sub_items)) and sub_items[idx + 1].get('nested')
                nl = '├─' if nxt else '└─'
                sub_detail = f'│   {nl} x{si["qty"]}'
                if si.get('detail'):
                    sub_detail += f'  {si["detail"]}'
            elif si.get('is_parent'):
                sub_detail = f'{tree} {si["sku"]} x{si["qty"]}'
                if si.get('detail'):
                    sub_detail += f'  {si["detail"]}'
                if si.get('batch_count'):
                    sub_detail += f' ({si["batch_count"]} batches)'
            else:
                sub_detail = f'{tree} {si["sku"]} x{si["qty"]}'
                if si.get('detail'):
                    sub_detail += f'  {si["detail"]}'

            sub_status = si.get('status', '')
            sub_err    = si.get('error', '')
            ws.append(['', '', '', sub_detail, sub_status, sub_err])
            sr = ws.max_row
            if sub_status == 'Skipped':
                ws.cell(sr, 4).fill = SUB_SKIP_FILL
            elif sub_err:
                ws.cell(sr, 4).fill = SUB_ERR_FILL
            else:
                ws.cell(sr, 4).fill = SUB_OK_FILL
            if sub_status in STATUS_FILL:
                ws.cell(sr, 5).fill = STATUS_FILL[sub_status]
            if sub_err:
                ws.cell(sr, 6).fill = ERR_CELL_FILL


# ── F1 Receiving ──────────────────────────────────────────────
def build_f1(wb):
    ws = wb.create_sheet('F1 Receiving')
    _cc_header(ws, 'F1', 'PO Receiving — Katana → WASP')

    _write_flow_event(ws, 'F1', 'WK-001', '2026-02-24 08:03', 'PO-1042',
        '3 items → RECEIVING-DOCK, PRODUCTION @ MMH Kelowna', 'Complete', sub_items=[
            {'sku': 'UFC-4OZ',     'qty': 200, 'detail': '→ RECEIVING-DOCK  lot:L2025-041', 'status': 'Added'},
            {'sku': 'B-WAX-500G',  'qty': 50,  'detail': '→ PRODUCTION  lot:L2025-039',     'status': 'Added'},
            {'sku': 'COLLAGEN-30', 'qty': 100, 'detail': '→ PRODUCTION  lot:L2025-040',     'status': 'Added'},
        ])

    _write_flow_event(ws, 'F1', 'WK-006', '2026-02-24 13:05', 'PO-1047',
        '2 items → RECEIVING-DOCK, PRODUCTION @ MMH Kelowna', 'Complete', sub_items=[
            {'sku': 'UFC-8OZ', 'qty': 500, 'detail': '→ RECEIVING-DOCK',
             'status': '', 'is_parent': True, 'batch_count': 2},
            {'sku': '', 'qty': 300, 'detail': 'lot:L2025-111  exp:2027-11-30', 'status': 'Added', 'nested': True},
            {'sku': '', 'qty': 200, 'detail': 'lot:L2025-112  exp:2027-12-31', 'status': 'Added', 'nested': True},
            {'sku': 'B-WAX-500G', 'qty': 40, 'detail': '→ PRODUCTION  lot:L2025-039', 'status': 'Added'},
        ])

    _write_flow_event(ws, 'F1', 'WK-011', '2026-02-25 08:55', 'PO-1049',
        '3 items → RECEIVING-DOCK @ MMH Kelowna  1 failed', 'Partial',
        error='1 item not in WASP', sub_items=[
            {'sku': 'UFC-4OZ', 'qty': 30, 'detail': '→ RECEIVING-DOCK  lot:L2025-041', 'status': 'Added'},
            {'sku': 'OP-195',  'qty': 50, 'detail': '→ RECEIVING-DOCK',                'status': 'Skipped',
             'error': 'Skipped prefix'},
            {'sku': 'UFC-8OZ', 'qty': 20, 'detail': '→ RECEIVING-DOCK',                'status': 'Failed',
             'error': 'Item not in WASP'},
        ])

    _write_flow_event(ws, 'F1', 'WK-017', '2026-02-26 08:10', 'PO-1053',
        '3 items → PRODUCTION, RECEIVING-DOCK @ MMH Kelowna', 'Complete', sub_items=[
            {'sku': 'COLLAGEN-30', 'qty': 200, 'detail': '→ PRODUCTION  lot:LOT-20260201', 'status': 'Added'},
            {'sku': 'B-WAX-500G',  'qty': 80,  'detail': '→ PRODUCTION  lot:L2026-003',    'status': 'Added'},
            {'sku': 'UFC-4OZ',     'qty': 60,  'detail': '→ RECEIVING-DOCK  lot:L2026-001','status': 'Added'},
        ])

    _write_flow_event(ws, 'F1', 'WK-019', '2026-02-26 14:30', 'PO-1055',
        '2 items → RECEIVING-DOCK @ Storage Warehouse', 'Complete', sub_items=[
            {'sku': 'UFC-4OZ', 'qty': 120, 'detail': '→ SW-STORAGE  lot:L2026-001', 'status': 'Added'},
            {'sku': 'UFC-8OZ', 'qty': 80,  'detail': '→ SW-STORAGE  lot:L2026-002', 'status': 'Added'},
        ])

    _write_flow_event(ws, 'F1', 'WK-020', '2026-02-26 16:00', 'PO-1056',
        '1 item → RECEIVING-DOCK @ MMH Kelowna', 'Partial', error='1 failed', sub_items=[
            {'sku': 'B-WAX-500G', 'qty': 75, 'detail': '→ RECEIVING-DOCK  lot:L2026-003', 'status': 'Added'},
            {'sku': 'OP-195',     'qty': 10, 'detail': '→ RECEIVING-DOCK',                 'status': 'Failed',
             'error': 'Item not in WASP'},
        ])

    return ws


# ── F2 Adjustments ───────────────────────────────────────────
def build_f2(wb):
    ws = wb.create_sheet('F2 Adjustments')
    _cc_header(ws, 'F2', 'Stock Adjustments — WASP → Katana')

    _write_flow_event(ws, 'F2', 'WK-002', '2026-02-24 09:17', 'SA',
        'UFC-8OZ x48  lot:L2025-112  WASP → Katana @ RECEIVING-DOCK @ MMH Kelowna',
        'Synced')

    _write_flow_event(ws, 'F2', 'WK-005', '2026-02-24 12:30', 'SA',
        '3 items  WASP → Katana @ PRODUCTION @ MMH Kelowna',
        'Partial', error='1 failed', sub_items=[
            {'sku': 'B-WAX-500G',  'qty': 5, 'detail': 'PRODUCTION  lot:L2025-039  remove', 'status': 'Synced'},
            {'sku': 'COLLAGEN-30', 'qty': 3, 'detail': 'PRODUCTION  lot:L2025-040  remove', 'status': 'Synced'},
            {'sku': 'OP-195',      'qty': 2, 'detail': 'PRODUCTION  remove',                'status': 'Failed',
             'error': 'Item not in Katana'},
        ])

    _write_flow_event(ws, 'F2', 'WK-009', '2026-02-24 16:02', 'SA',
        'UFC-4OZ x10  WASP → Katana @ SW-STORAGE @ Storage Warehouse',
        'Skipped')

    _write_flow_event(ws, 'F2', 'WK-012', '2026-02-25 09:30', 'SA',
        '2 items  WASP → Katana @ SW-STORAGE @ Storage Warehouse',
        'Synced', sub_items=[
            {'sku': 'UFC-4OZ', 'qty': 24, 'detail': 'SW-STORAGE  lot:L2025-041  add', 'status': 'Synced'},
            {'sku': 'UFC-8OZ', 'qty': 12, 'detail': 'SW-STORAGE  lot:L2025-112  add', 'status': 'Synced'},
        ])

    _write_flow_event(ws, 'F2', 'WK-016', '2026-02-25 15:05', 'SA',
        'UFC-4OZ x5  Site not mapped  WASP → Katana',
        'Failed', error='Site not mapped')

    _write_flow_event(ws, 'F2', 'WK-021', '2026-02-26 09:00', 'SA',
        'COLLAGEN-30 x30  lot:LOT-20260201  WASP → Katana @ PRODUCTION @ MMH Kelowna',
        'Synced')

    _write_flow_event(ws, 'F2', 'WK-022', '2026-02-26 11:15', 'SA',
        '4 items  WASP → Katana @ UNSORTED @ MMH Kelowna',
        'Failed', error='Location not found in WASP', sub_items=[
            {'sku': 'UFC-4OZ',    'qty': 8,  'detail': 'UNSORTED  add', 'status': 'Failed', 'error': 'Location not found in WASP'},
            {'sku': 'UFC-8OZ',    'qty': 4,  'detail': 'UNSORTED  add', 'status': 'Failed', 'error': 'Location not found in WASP'},
            {'sku': 'B-WAX-500G', 'qty': 6,  'detail': 'UNSORTED  add', 'status': 'Failed', 'error': 'Location not found in WASP'},
            {'sku': 'COLLAGEN-30','qty': 10, 'detail': 'UNSORTED  add', 'status': 'Failed', 'error': 'Location not found in WASP'},
        ])

    return ws


# ── F3 Transfers ─────────────────────────────────────────────
def build_f3(wb):
    ws = wb.create_sheet('F3 Transfers')
    _cc_header(ws, 'F3', 'Amazon/Transfer Orders — Katana → WASP')

    _write_flow_event(ws, 'F3', 'WK-007', '2026-02-24 14:20', 'SO-88231',
        '1 item → SHOPIFY', 'Complete', sub_items=[
            {'sku': 'B-WAX-500G', 'qty': 20, 'detail': '→ SHOPIFY', 'status': 'Complete'},
        ])

    _write_flow_event(ws, 'F3', 'WK-014', '2026-02-25 11:40', 'SO-88290',
        '2 items → SHOPIFY', 'Complete', sub_items=[
            {'sku': 'UFC-4OZ',    'qty': 15, 'detail': '→ SHOPIFY', 'status': 'Complete'},
            {'sku': 'COLLAGEN-30','qty': 10, 'detail': '→ SHOPIFY', 'status': 'Complete'},
        ])

    _write_flow_event(ws, 'F3', 'WK-023', '2026-02-25 16:10', 'SO-88305',
        '3 items → SHOPIFY', 'Partial', error='1 failed', sub_items=[
            {'sku': 'UFC-4OZ',    'qty': 6,  'detail': '→ SHOPIFY', 'status': 'Complete'},
            {'sku': 'UFC-8OZ',    'qty': 4,  'detail': '→ SHOPIFY', 'status': 'Complete'},
            {'sku': 'B-WAX-500G', 'qty': 2,  'detail': '→ SHOPIFY', 'status': 'Failed',
             'error': 'Insufficient qty at location'},
        ])

    _write_flow_event(ws, 'F3', 'WK-024', '2026-02-26 07:50', 'SO-88320',
        '1 item → SHOPIFY', 'Complete', sub_items=[
            {'sku': 'UFC-8OZ', 'qty': 24, 'detail': '→ SHOPIFY', 'status': 'Complete'},
        ])

    _write_flow_event(ws, 'F3', 'WK-025', '2026-02-26 10:05', 'SO-88335',
        '2 items → SHOPIFY', 'Failed', error='Location not found in WASP', sub_items=[
            {'sku': 'COLLAGEN-30', 'qty': 5,  'detail': '→ SHOPIFY', 'status': 'Failed',
             'error': 'Location not found in WASP'},
            {'sku': 'B-WAX-500G',  'qty': 3,  'detail': '→ SHOPIFY', 'status': 'Failed',
             'error': 'Location not found in WASP'},
        ])

    _write_flow_event(ws, 'F3', 'WK-026', '2026-02-26 13:22', 'SO-88351',
        '1 item → SHOPIFY', 'Complete', sub_items=[
            {'sku': 'UFC-4OZ', 'qty': 10, 'detail': '→ SHOPIFY', 'status': 'Complete'},
        ])

    return ws


# ── F4 Manufacturing ─────────────────────────────────────────
def build_f4(wb):
    ws = wb.create_sheet('F4 Manufacturing')
    _cc_header(ws, 'F4', 'MO Complete — Katana → WASP')

    _write_flow_event(ws, 'F4', 'WK-003', '2026-02-24 10:45', 'MO-7246',
        '3 items → PROD-RECEIVING @ MMH Kelowna', 'Complete', sub_items=[
            {'sku': 'B-WAX-500G',  'qty': 12,  'detail': 'PRODUCTION  lot:L2025-039  remove', 'status': 'Added'},
            {'sku': 'COLLAGEN-30', 'qty': 8,   'detail': 'PRODUCTION  lot:L2025-040  remove', 'status': 'Added'},
            {'sku': 'UFC-4OZ',     'qty': 240, 'detail': '→ PROD-RECEIVING  lot:UFC410B  exp:2029-02-28', 'status': 'Added'},
        ])

    _write_flow_event(ws, 'F4', 'WK-008', '2026-02-24 15:10', 'MO-7251',
        '2 items → PROD-RECEIVING @ MMH Kelowna', 'Failed',
        error='Insufficient qty at location', sub_items=[
            {'sku': 'B-WAX-500G',  'qty': 6,   'detail': 'PRODUCTION  remove', 'status': 'Failed',
             'error': 'Insufficient qty at location'},
            {'sku': 'COLLAGEN-30', 'qty': 120, 'detail': '→ PROD-RECEIVING',   'status': 'Failed',
             'error': 'Insufficient qty at location'},
        ])

    _write_flow_event(ws, 'F4', 'WK-013', '2026-02-25 10:15', 'MO-7260',
        '3 items → PROD-RECEIVING @ MMH Kelowna', 'Partial',
        error='1 failed', sub_items=[
            {'sku': 'B-WAX-500G',  'qty': 9,   'detail': 'PRODUCTION  lot:L2025-039  remove', 'status': 'Added'},
            {'sku': 'COLLAGEN-30', 'qty': 6,   'detail': 'PRODUCTION  remove',                'status': 'Failed',
             'error': 'Lot not found at location'},
            {'sku': 'UFC-8OZ',     'qty': 180, 'detail': '→ PROD-RECEIVING  lot:UFC412A  exp:2028-12-31', 'status': 'Added'},
        ])

    _write_flow_event(ws, 'F4', 'WK-018', '2026-02-26 09:45', 'MO-7271',
        '4 items → PROD-RECEIVING @ MMH Kelowna', 'Complete', sub_items=[
            {'sku': 'B-WAX-500G',  'qty': 18,  'detail': 'PRODUCTION  lot:L2025-039  remove', 'status': 'Added'},
            {'sku': 'COLLAGEN-30', 'qty': 12,  'detail': 'PRODUCTION  lot:L2025-040  remove', 'status': 'Added'},
            {'sku': 'OP-195',      'qty': 2,   'detail': 'PRODUCTION  remove',                'status': 'Skipped'},
            {'sku': 'UFC-4OZ',     'qty': 360, 'detail': '→ PROD-RECEIVING  lot:UFC413C  exp:2029-02-28', 'status': 'Added'},
        ])

    _write_flow_event(ws, 'F4', 'WK-027', '2026-02-26 15:30', 'MO-7280',
        '2 items → PROD-RECEIVING @ MMH Kelowna', 'Complete', sub_items=[
            {'sku': 'B-WAX-500G',  'qty': 5,   'detail': 'PRODUCTION  lot:L2026-003  remove', 'status': 'Added'},
            {'sku': 'UFC-4OZ',     'qty': 100, 'detail': '→ PROD-RECEIVING  lot:UFC414D  exp:2029-03-31', 'status': 'Added'},
        ])

    return ws


# ── F5 Shipping ───────────────────────────────────────────────
def build_f5(wb):
    ws = wb.create_sheet('F5 Shipping')
    _cc_header(ws, 'F5', 'Order Fulfillment — Katana → WASP / ShipStation')

    _write_flow_event(ws, 'F5', 'WK-004', '2026-02-24 11:02', '#90412',
        '2 items → SHIPPING-DOCK', 'Complete', sub_items=[
            {'sku': 'UFC-4OZ', 'qty': 2, 'detail': '→ SHIPPING-DOCK', 'status': 'Picked'},
            {'sku': 'UFC-8OZ', 'qty': 1, 'detail': '→ SHIPPING-DOCK', 'status': 'Picked'},
        ])

    _write_flow_event(ws, 'F5', 'WK-010', '2026-02-24 17:45', '#90589',
        '4 items → SHIPPING-DOCK', 'Complete', sub_items=[
            {'sku': 'UFC-4OZ',    'qty': 6, 'detail': '→ SHIPPING-DOCK', 'status': 'Picked'},
            {'sku': 'UFC-8OZ',    'qty': 3, 'detail': '→ SHIPPING-DOCK', 'status': 'Picked'},
            {'sku': 'B-WAX-500G', 'qty': 2, 'detail': '→ SHIPPING-DOCK', 'status': 'Picked'},
            {'sku': 'COLLAGEN-30','qty': 1, 'detail': '→ SHIPPING-DOCK', 'status': 'Picked'},
        ])

    _write_flow_event(ws, 'F5', 'WK-015', '2026-02-25 14:22', '#90601',
        '1 item  Duplicate order number', 'Failed', error='Duplicate order number')

    _write_flow_event(ws, 'F5', 'WK-028', '2026-02-25 16:55', '#90615',
        '3 items → SHIPPING-DOCK', 'Partial', error='1 failed', sub_items=[
            {'sku': 'UFC-4OZ',    'qty': 4, 'detail': '→ SHIPPING-DOCK', 'status': 'Picked'},
            {'sku': 'UFC-8OZ',    'qty': 2, 'detail': '→ SHIPPING-DOCK', 'status': 'Picked'},
            {'sku': 'B-WAX-500G', 'qty': 1, 'detail': '→ SHIPPING-DOCK', 'status': 'Failed',
             'error': 'Insufficient qty at location'},
        ])

    _write_flow_event(ws, 'F5', 'WK-029', '2026-02-26 10:30', '#90630',
        '1 item → SHIPPING-DOCK', 'Complete', sub_items=[
            {'sku': 'COLLAGEN-30', 'qty': 5, 'detail': '→ SHIPPING-DOCK', 'status': 'Picked'},
        ])

    _write_flow_event(ws, 'F5', 'WK-030', '2026-02-26 14:10', '#90648',
        '2 items → SHIPPING-DOCK', 'Complete', sub_items=[
            {'sku': 'UFC-4OZ', 'qty': 3, 'detail': '→ SHIPPING-DOCK', 'status': 'Picked'},
            {'sku': 'UFC-8OZ', 'qty': 2, 'detail': '→ SHIPPING-DOCK', 'status': 'Picked'},
        ])

    return ws


# ═══════════════════════════════════════════════════════════════
# TAB 7 – Adjustments
# Row 1: Title, Row 2: Instructions, Row 3: Headers (frozen)
# Columns: Action | SKU | Qty | Site | Location | Lot | Expiry | Notes | Status | Error
# ═══════════════════════════════════════════════════════════════

def build_adjustments(wb):
    ws = wb.create_sheet('Adjustments')

    # Row 1 – title
    ws.append(['STOCK ADJUSTMENTS', '', '', '', '', '', '', '', '', ''])
    ws['A1'].font = Font(bold=True, size=12)

    # Row 2 – instructions
    ws.append([
        'Enter rows below, set Status to PENDING, then run processAdjustments()',
        '', '', '', '', '', '', '', '', ''
    ])
    ws['A2'].font = GREY_FONT

    # Row 3 – headers
    headers = ['Action', 'SKU', 'Qty', 'Site', 'Location', 'Lot', 'Expiry', 'Notes', 'Status', 'Error']
    ws.append(headers)
    for c, cell in enumerate(ws[3], start=1):
        cell.font = BOLD
        cell.fill = HEADER_COL_FILL

    ws.freeze_panes = 'A4'

    # Column widths (mirrors GAS sheet.setColumnWidth calls)
    col_widths = [10, 18, 7, 18, 16, 18, 12, 40, 10, 38]
    for idx, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    # Helper to write a data row
    STATUS_ADJ = {
        'PENDING':  PatternFill('solid', fgColor='FFF8E1'),
        'Complete': PatternFill('solid', fgColor='D4EDDA'),
        'Failed':   PatternFill('solid', fgColor='F8D7DA'),
        'Skip':     PatternFill('solid', fgColor='E0E0E0'),
    }

    def adj_row(action, sku, qty, site, location, lot, expiry, notes, status, error=''):
        ws.append([action, sku, qty, site, location, lot, expiry, notes, status, error])
        r = ws.max_row
        ws.cell(r, 9).fill = STATUS_ADJ.get(status, PatternFill())
        if error:
            ws.cell(r, 10).fill = ERR_CELL_FILL
        # Bold action
        ws.cell(r, 1).font = Font(bold=True)

    # 8 realistic adjustment rows
    adj_row('ADD',    'UFC-4OZ',     50,  'MMH Kelowna',       'RECEIVING-DOCK',
            'L2025-041', '2027-12-31',
            'Emergency restock from supplier overflow', 'Complete')

    adj_row('REMOVE', 'B-WAX-500G',  10,  'MMH Kelowna',       'PRODUCTION',
            'L2025-039', '2027-06-30',
            'Expired units — quarantine removal', 'Complete')

    adj_row('ADJUST', 'COLLAGEN-30', -5,  'MMH Kelowna',       'PRODUCTION',
            'L2025-040', '2027-09-30',
            'Cycle count correction — over-count on shelf', 'Complete')

    adj_row('ADD',    'UFC-8OZ',     25,  'Storage Warehouse', 'SW-STORAGE',
            'L2025-112', '2027-12-31',
            'Transfer from off-site warehouse', 'PENDING')

    adj_row('REMOVE', 'OP-195',      3,   'MMH Kelowna',       'PRODUCTION',
            '', '',
            'Write-off — damaged during production', 'Failed',
            'Item not in WASP')

    adj_row('ADD',    'B-WAX-500G',  15,  'MMH Kelowna',       'PROD-RECEIVING',
            'LOT-20260201', '2028-02-28',
            'Manual receipt — supplier short-shipped PO-1049', 'PENDING')

    adj_row('ADJUST', 'UFC-4OZ',     10,  'Storage Warehouse', 'SW-STORAGE',
            'L2026-001', '2028-01-31',
            'Reconcile after physical count — found extra units', 'Complete')

    adj_row('REMOVE', 'COLLAGEN-30', 8,   'MMH Kelowna',       'SHOPIFY',
            'LOT-20250115', '2027-03-31',
            'Recalled batch — withdraw from SHOPIFY location', 'PENDING')

    return ws


# ═══════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════

def main():
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Build tabs in order
    ws_activity    = build_activity(wb)
    ws_f1          = build_f1(wb)
    ws_f2          = build_f2(wb)
    ws_f3          = build_f3(wb)
    ws_f4          = build_f4(wb)
    ws_f5          = build_f5(wb)
    ws_adjustments = build_adjustments(wb)

    out_path = r'C:\Users\demch\Downloads\wasp-katana\tracker-mock.xlsx'
    wb.save(out_path)

    # Print summary
    print(f'Saved: {out_path}')
    print()
    print('Tabs created:')
    for ws in [ws_activity, ws_f1, ws_f2, ws_f3, ws_f4, ws_f5, ws_adjustments]:
        data_rows = ws.max_row - 3  # subtract 3-row CC header
        print(f'  {ws.title:<22}  {ws.max_row} total rows  ({data_rows} data rows after header)')

    # Count F1 parent rows and child rows in Activity tab
    print()
    print('F1 Receiving fix summary (Activity tab):')
    f1_parents = 0
    f1_children = 0
    in_f1_event = False
    for row in ws_activity.iter_rows(min_row=4, values_only=True):
        exec_id, ts, flow, details, status = row[0], row[1], row[2], row[3], row[4]
        if exec_id and flow == 'F1 Receiving':
            f1_parents += 1
            in_f1_event = True
        elif not exec_id and in_f1_event and details and ('├─' in str(details) or '└─' in str(details)):
            f1_children += 1
        elif exec_id:
            in_f1_event = False
    print(f'  F1 parent rows updated : {f1_parents}')
    print(f'  F1 child rows added    : {f1_children}')

    print()
    print('F1 Receiving fix summary (F1 tab):')
    f1t_parents = 0
    f1t_children = 0
    for row in ws_f1.iter_rows(min_row=4, values_only=True):
        exec_id, details = row[0], row[3]
        if exec_id:
            f1t_parents += 1
        elif details and ('├─' in str(details) or '└─' in str(details)):
            f1t_children += 1
    print(f'  F1 tab parent events   : {f1t_parents}')
    print(f'  F1 tab child rows      : {f1t_children}')


if __name__ == '__main__':
    main()
