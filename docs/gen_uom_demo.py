import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

wb = openpyxl.Workbook()

# colours
F4_HDR  = "f0e6f6"
F1_HDR  = "cce5ff"
F2_HDR  = "b2dfdb"
SUB_OK  = "f1faf1"
SUB_SKIP= "fff9e6"
SUB_PROD= "e8f5e9"
WHITE   = "ffffff"
GREY    = "888888"
BLU     = "1a56db"
GRN     = "1a7340"
ORG     = "b45309"

def fill(h): return PatternFill("solid", fgColor=h)
def fnt(bold=False, color="000000", size=9): return Font(bold=bold, color=color, size=size, name="Calibri")
def bot_border(): return Border(bottom=Side(style="thin", color="e0e0e0"))

def write_header(ws, row, wk, ts, flow, details, status, note, bg):
    for c, v in enumerate([wk, ts, flow, details, status, note], 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.fill = fill(bg)
        cell.border = bot_border()
        cell.alignment = Alignment(vertical="center")
        if c == 1:   cell.font = fnt(bold=True, color="333333")
        elif c == 4: cell.font = fnt(bold=True, color=BLU, size=9)
        elif c == 5: cell.font = fnt(bold=True, color="444444")
        else:        cell.font = fnt(color="555555")

def write_sub(ws, row, details, status, note="", bg=WHITE, det_col="000000", st_col="1a7340"):
    for c in range(1, 7):
        ws.cell(row=row, column=c).fill = fill(bg)
        ws.cell(row=row, column=c).border = bot_border()
    ws.cell(row=row, column=4, value=details).font = fnt(color=det_col)
    ws.cell(row=row, column=4).alignment = Alignment(indent=1)
    ws.cell(row=row, column=5, value=status).font = fnt(bold=True, color=st_col)
    ws.cell(row=row, column=6, value=note).font = fnt(color=GREY)

def setup_sheet(ws):
    for col, w in [(1,9),(2,16),(3,14),(4,82),(5,12),(6,24)]:
        ws.column_dimensions[get_column_letter(col)].width = w
    for c, h in enumerate(["ID","Time","Flow","Details","Status","Error/Note"], 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.fill = fill("e8eaf6")
        cell.font = fnt(bold=True, color="222222")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A4"

# ── Sheet 1: Current (no UOM) ──────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Current (no UOM)"
setup_sheet(ws1)

r = 4
write_header(ws1, r, "WK-043","2026-03-02 13:50","F4 Manufacturing",
    "MO-7286 (UFC414C) // MAR 02  MS-IP-4 x478  UFC414C -> PRODUCTION",
    "Complete","1 skipped",F4_HDR)
r+=1
write_sub(ws1,r,"  |- LABSSEAL-4 x478  PRODUCTION","Consumed",bg=SUB_OK)
r+=1
write_sub(ws1,r,"  |- EO-LAV x300  PRODUCTION  lot:5005828385  exp:2029-07-01","Consumed",bg=SUB_OK)
r+=1
write_sub(ws1,r,"  |- G-OIL x2.045  PRODUCTION  lot:06603/25  exp:2028-06-30","Consumed",bg=SUB_OK)
r+=1
write_sub(ws1,r,"  |- B-WAX x7200  PRODUCTION  lot:MHF121525  exp:2026-12-15","Consumed",bg=SUB_OK)
r+=1
write_sub(ws1,r,"  |- O-OIL x38.249  PRODUCTION  lot:47/2025  exp:2028-09-30","Consumed",bg=SUB_OK)
r+=1
write_sub(ws1,r,"  |- IP-EGG x3250  PRODUCTION","Skipped","Not in WASP",SUB_SKIP,"000000",ORG)
r+=1
write_sub(ws1,r,"  L- MS-IP-4 x478  PROD-RECEIVING  lot:UFC414C  exp:2029-03-02","Produced",bg=SUB_PROD,det_col=GRN,st_col=GRN)
r+=1

write_header(ws1,r,"WK-046","2026-03-02 14:04","F1 Receiving",
    "PO-569  L-M x10  -> RECEIVING-DOCK @ MMH Kelowna","Received","",F1_HDR)
r+=1
write_sub(ws1,r,"  L- L-M x10  -> RECEIVING-DOCK","Added",bg=SUB_PROD,det_col=GRN,st_col=GRN)
r+=1

write_header(ws1,r,"WK-047","2026-03-02 14:05","F2 Adjustments",
    "WASP Adjustment  EO-LAV x500","Synced","",F2_HDR)
r+=1
write_sub(ws1,r,"  L- EO-LAV x500  PRODUCTION  remove","Synced",bg=SUB_OK)
r+=1

# ── Sheet 2: Proposed (with UOM) ──────────────────────────────────────────
ws2 = wb.create_sheet("Proposed (with UOM)")
setup_sheet(ws2)

r = 4
write_header(ws2,r,"WK-043","2026-03-02 13:50","F4 Manufacturing",
    "MO-7286 (UFC414C) // MAR 02  MS-IP-4 x478 ea  UFC414C -> PRODUCTION",
    "Complete","1 skipped",F4_HDR)
r+=1
write_sub(ws2,r,"  |- LABSSEAL-4 x478 mL  PRODUCTION","Consumed",bg=SUB_OK)
r+=1
write_sub(ws2,r,"  |- EO-LAV x300 mL  PRODUCTION  lot:5005828385  exp:2029-07-01","Consumed",bg=SUB_OK)
r+=1
write_sub(ws2,r,"  |- G-OIL x2.045 kg  PRODUCTION  lot:06603/25  exp:2028-06-30","Consumed",bg=SUB_OK)
r+=1
write_sub(ws2,r,"  |- B-WAX x7200 g  PRODUCTION  lot:MHF121525  exp:2026-12-15","Consumed",bg=SUB_OK)
r+=1
write_sub(ws2,r,"  |- O-OIL x38.249 mL  PRODUCTION  lot:47/2025  exp:2028-09-30","Consumed",bg=SUB_OK)
r+=1
write_sub(ws2,r,"  |- IP-EGG x3250 g  PRODUCTION","Skipped","Not in WASP",SUB_SKIP,"000000",ORG)
r+=1
write_sub(ws2,r,"  L- MS-IP-4 x478 ea  PROD-RECEIVING  lot:UFC414C  exp:2029-03-02","Produced",bg=SUB_PROD,det_col=GRN,st_col=GRN)
r+=1

# F1 with purchase UOM conversion note
write_header(ws2,r,"WK-046","2026-03-02 14:04","F1 Receiving",
    "PO-569  L-M x10 dozen -> 120 pcs  RECEIVING-DOCK @ MMH Kelowna","Received","",F1_HDR)
r+=1
write_sub(ws2,r,"  L- L-M x120 pcs  -> RECEIVING-DOCK  (10 dozen x12)","Added",bg=SUB_PROD,det_col=GRN,st_col=GRN)
r+=1

# F2 with UOM
write_header(ws2,r,"WK-047","2026-03-02 14:05","F2 Adjustments",
    "WASP Adjustment  EO-LAV x500 mL","Synced","",F2_HDR)
r+=1
write_sub(ws2,r,"  L- EO-LAV x500 mL  PRODUCTION  remove","Synced",bg=SUB_OK)
r+=1

# notes
r+=1
ws2.cell(row=r,column=1,value="How it works:").font = fnt(bold=True,color="222222",size=10)
r+=1
notes = [
    "Format:  x{qty} {uom}   e.g.  x300 mL  or  x7200 g  or  x478 ea",
    "UOM is placed immediately after the quantity, before the location.",
    "",
    "Where each flow gets its UOM:",
    "  F4  ingredients + output  ->  variant.uom already fetched per item (ingUom already in results)",
    "  F1  received rows         ->  variant.uom from variant lookup; purchase UOM shown in header if conversion rate > 1",
    "  F2  WASP adjustments      ->  variant.uom from the Katana variant lookup done in processSiteBatchAdd/Remove",
    "  F3 / F5 / F6              ->  variant.uom from variant lookup (already resolved when building sub-items)",
    "",
    "If variant.uom is empty in Katana the field is simply omitted — no change to current display.",
    "No extra API calls are needed; UOM comes from the variant object already fetched in every flow.",
]
for note in notes:
    cell = ws2.cell(row=r, column=1, value=note)
    cell.font = fnt(color="444444", size=9)
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r+=1

out = r"C:\Users\Admin\Documents\claude-projects\wasp-katana\docs\uom-demo.xlsx"
os.makedirs(os.path.dirname(out), exist_ok=True)
wb.save(out)
print("Saved:", out)
