import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
import os

BASE_DIR = r"C:\Users\demch\Downloads\wasp-katana"

CANDIDATES = [
    os.path.join(BASE_DIR, "tracker-mock-new.xlsx"),
    os.path.join(BASE_DIR, "tracker-mock.xlsx"),
]
FALLBACK  = os.path.join(BASE_DIR, "tracker-mock-adj-demo.xlsx")

# ---------------------------------------------------------------------------
# Helper: thin border factory
# ---------------------------------------------------------------------------
thin = Side(style="thin", color="D0D0D0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

# ---------------------------------------------------------------------------
# Colors
# ---------------------------------------------------------------------------
HEADER_BG   = "1F2937"
HEADER_FG   = "FFFFFF"

ROW_ADD     = "E8F5E9"
ROW_REMOVE  = "FFEBEE"
ROW_MOVE    = "E3F2FD"
ROW_SYNC    = "FFF8E1"

OK_BG       = "C8E6C9"
ERR_BG      = "FFCDD2"

POS_FG      = "2E7D32"
POS_BG      = "C8E6C9"
NEG_FG      = "C62828"
NEG_BG      = "FFCDD2"

LINK_FG     = "1565C0"

# ---------------------------------------------------------------------------
# Column definitions: (header, width)
# ---------------------------------------------------------------------------
COLUMNS = [
    ("Timestamp",    22),
    ("Source",       14),
    ("Action",       11),
    ("User",         16),
    ("SKU",          16),
    ("Item Name",    24),
    ("Site",         18),
    ("Location",     18),
    ("Lot / Batch",  18),
    ("Expiry",       13),
    ("Diff",         10),
    ("Katana SA#",   15),
    ("Status",       10),
    ("Note",         32),
]

# ---------------------------------------------------------------------------
# Demo data rows  (14 fields each, matching columns A-N)
# Diff values stored as numeric where possible so we can apply number format
# ---------------------------------------------------------------------------
ROWS = [
    # Row 1
    ("2026-02-26 14:23:05", "WASP",       "Remove", "john.k",  "B-WAX",     "Beeswax",                 "MMH Kelowna", "PRODUCTION",           "MHF121525",    "2026-12-14", -200, "B-WAX",     "OK",    ""),
    # Row 2
    ("2026-02-26 14:23:07", "WASP",       "Remove", "john.k",  "EO-LAV",    "Essential Oil Lavender",  "MMH Kelowna", "PRODUCTION",           "EO-2025-11",   "2025-11-30",  -50, "EO-LAV +2", "OK",    ""),
    # Row 3
    ("2026-02-26 14:23:07", "WASP",       "Remove", "john.k",  "EO-PEP",    "Essential Oil Peppermint","MMH Kelowna", "PRODUCTION",           "EO-2025-09",   "2025-09-15",  -30, "EO-LAV +2", "OK",    ""),
    # Row 4
    ("2026-02-26 14:23:07", "WASP",       "Remove", "john.k",  "EO-TT",     "Essential Oil Tea Tree",  "MMH Kelowna", "PRODUCTION",           "EO-2026-01",   "2026-01-20",  -75, "EO-LAV +2", "OK",    ""),
    # Row 5
    ("2026-02-26 15:04:22", "WASP",       "Add",    "sarah.m", "EGG-X",     "Egg",                     "MMH Kelowna", "PRODUCTION",           "PO-1055",      "",            +12, "EGG-X",     "OK",    ""),
    # Row 6
    ("2026-02-26 15:11:03", "WASP",       "Move",   "sarah.m", "UFC-4OZ",   "UFC 4oz",                 "MMH Kelowna", "RECEIVING-DOCK\u2192PRODUCTION", "CAR023", "",        60,  "",          "OK",    ""),
    # Row 7
    ("2026-02-26 15:45:18", "WASP",       "Remove", "john.k",  "LCP-2",     "LCP Cream 2oz",           "MMH Kelowna", "PROD-RECEIVING",       "CAR031",       "",            -10, "LCP-2",     "ERROR", "Batch tracking required"),
    # Row 8
    ("2026-02-26 16:02:44", "Sync Sheet", "Add",    "",        "INFBAG-160","Infusion Bag",            "MMH Kelowna", "PRODUCTION",           "INF-2025-10",  "2025-10-31", +500, "",          "OK",    ""),
    # Row 9
    ("2026-02-26 16:03:11", "Sync Sheet", "Remove", "",        "B-PROP",    "Propolis",                "MMH Kelowna", "PRODUCTION",           "PROP-2026-02", "2026-02-28", -100, "",          "OK",    ""),
    # Row 10
    ("2026-02-27 08:45:00", "WASP",       "Add",    "ed.d",    "EGG-X",     "Egg",                     "MMH Kelowna", "PRODUCTION",           "PO-1061",      "",            +12, "EGG-X",     "OK",    "uom:1dozen\u219212pcs"),
    # Row 11
    ("2026-02-27 09:12:33", "WASP",       "Remove", "sarah.m", "UFC-4OZ",   "UFC 4oz",                 "MMH Kelowna", "PROD-RECEIVING",       "L2026-003",    "",             -2, "UFC-4OZ",   "OK",    "Retention sample"),
    # Row 12
    ("2026-02-27 09:30:15", "WASP",       "Add",    "john.k",  "NEWSKU-1",  "(unknown)",               "MMH Kelowna", "PRODUCTION",           "",             "",            +10, "",          "ERROR", "Item not in Katana"),
]

ACTION_ROW_COLOR = {
    "Add":    ROW_ADD,
    "Remove": ROW_REMOVE,
    "Move":   ROW_MOVE,
    "Sync":   ROW_SYNC,
}

# ---------------------------------------------------------------------------
# Build the sheet
# ---------------------------------------------------------------------------
def build_sheet(ws):
    # --- Header row ---
    header_fill = make_fill(HEADER_BG)
    header_font = Font(name="Calibri", bold=True, color=HEADER_FG, size=10)

    for col_idx, (header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 18

    # Freeze row 1
    ws.freeze_panes = "A2"

    # --- Data rows ---
    for row_idx, row_data in enumerate(ROWS, start=2):
        action     = row_data[2]   # col C  (0-based index 2)
        diff_val   = row_data[10]  # col K  (0-based index 10)
        sa_val     = row_data[11]  # col L  (0-based index 11)
        status_val = row_data[12]  # col M  (0-based index 12)

        row_bg_hex = ACTION_ROW_COLOR.get(action, ROW_ADD)
        row_fill   = make_fill(row_bg_hex)

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border    = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)

            # Default: row background
            cell.fill = row_fill
            cell.font = Font(name="Calibri", size=10)

            # ---- Column K (Diff) ----
            if col_idx == 11:  # K
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if isinstance(value, (int, float)):
                    if value > 0:
                        cell.fill = make_fill(POS_BG)
                        cell.font = Font(name="Calibri", size=10, color=POS_FG, bold=True)
                        # Show + prefix as text so it's explicit
                        cell.value = value   # numeric, formatted below
                        cell.number_format = r'"+"\#,##0;"-"\#,##0'
                    elif value < 0:
                        cell.fill = make_fill(NEG_BG)
                        cell.font = Font(name="Calibri", size=10, color=NEG_FG, bold=True)
                        cell.number_format = r'"+"\#,##0;"-"\#,##0'
                    else:
                        cell.number_format = "General"

            # ---- Column L (Katana SA#) ----
            elif col_idx == 12:  # L
                if value and str(value).strip():
                    cell.font = Font(name="Calibri", size=10, color=LINK_FG, bold=True)

            # ---- Column M (Status) ----
            elif col_idx == 13:  # M
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if value == "OK":
                    cell.fill = make_fill(OK_BG)
                    cell.font = Font(name="Calibri", size=10, bold=True, color="1B5E20")
                elif value == "ERROR":
                    cell.fill = make_fill(ERR_BG)
                    cell.font = Font(name="Calibri", size=10, bold=True, color="B71C1C")

        ws.row_dimensions[row_idx].height = 16

# ---------------------------------------------------------------------------
# Try to open / save
# ---------------------------------------------------------------------------
def try_save(path, wb):
    try:
        wb.save(path)
        print("Saved: " + path)
        return True
    except PermissionError:
        print("Locked (PermissionError): " + path)
        return False
    except Exception as ex:
        print("Error saving " + path + ": " + str(ex))
        return False

saved_path = None
for candidate in CANDIDATES:
    if not os.path.exists(candidate):
        print("Not found: " + candidate)
        continue
    print("Opening: " + candidate)
    try:
        wb = openpyxl.load_workbook(candidate)
    except Exception as ex:
        print("Cannot open " + candidate + ": " + str(ex))
        continue

    # Remove existing sheet if present, then re-add
    SHEET_NAME = "Adjustments Log"
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME)
    build_sheet(ws)

    if try_save(candidate, wb):
        saved_path = candidate
        break

if not saved_path:
    # Fallback: create fresh workbook with just the sheet
    print("Trying fallback: " + FALLBACK)
    # Load tracker-mock.xlsx as base if available
    base = CANDIDATES[-1] if os.path.exists(CANDIDATES[-1]) else None
    if base:
        try:
            wb = openpyxl.load_workbook(base)
        except Exception:
            wb = openpyxl.Workbook()
    else:
        wb = openpyxl.Workbook()

    SHEET_NAME = "Adjustments Log"
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME)
    build_sheet(ws)

    if try_save(FALLBACK, wb):
        saved_path = FALLBACK

if saved_path:
    print("\nDone. Adjustments Log sheet written to: " + saved_path)
else:
    print("\nFailed to save to any target.")
